from app.models import Person, Contact, parent_child_table, Employee, Subject, Subscription, Lesson, SchoolClass, \
    SubjectType, SubscriptionType, StudentAttendance, SchoolLessonJournal, Finance, UserAction
from datetime import datetime, timedelta
import pytz
from app import db
from sqlalchemy import and_, or_
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, GradientFill
from openpyxl.utils import get_column_letter

DAYS_OF_WEEK = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
MONTHS = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль",
          "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
OPERATION_TYPES = {"cash": "нал", "bank": "счет", "balance": "депозит"}
CHILD = "Ребенок"
ADULT = "Взрослый"
TEACHER = "Учитель"
CHILD_SELF = "Сам ребенок"
CHOOSE = "Выбрать"
OTHER = "Другое"
LOCAL_TZ = pytz.timezone('Asia/Tbilisi')


def get_today_date():
    return datetime.now(LOCAL_TZ).date()


def conjugate_lessons(number):
    last_digit = number % 10
    last_two_digits = number % 100

    if 10 <= last_two_digits <= 20:
        return f"{number} занятий"
    elif last_digit == 1:
        return f"{number} занятие"
    elif 2 <= last_digit <= 4:
        return f"{number} занятия"
    else:
        return f"{number} занятий"


def conjugate_years(number):
    last_digit = number % 10
    last_two_digits = number % 100

    if 10 <= last_two_digits <= 20:
        return f"{number} лет"
    elif last_digit == 1:
        return f"{number} год"
    elif 2 <= last_digit <= 4:
        return f"{number} года"
    else:
        return f"{number} лет"


def conjugate_hours(number):
    last_digit = number % 10
    last_two_digits = number % 100

    if 10 <= last_two_digits <= 20:
        return f"{number} часов"
    elif last_digit == 1:
        return f"{number} час"
    elif 2 <= last_digit <= 4:
        return f"{number} часа"
    else:
        return f"{number} часов"


def conjugate_days(number):
    last_digit = number % 10
    last_two_digits = number % 100

    if 10 <= last_two_digits <= 20:
        return f"{number} дней"
    elif last_digit == 1:
        return f"{number} день"
    elif 2 <= last_digit <= 4:
        return f"{number} дня"
    else:
        return f"{number} дней"


def person_age(dob):
    today = get_today_date()
    age = get_today_date().year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    return conjugate_years(age)


def create_student(form, student_type):
    last_name = form.last_name.data
    first_name = form.first_name.data
    patronym = form.patronym.data
    status = form.status.data
    same_person = Person.query.filter_by(last_name=last_name, first_name=first_name).all()
    if not same_person:
        if student_type == 'child':
            dob = datetime.strptime(form.dob.data, '%d.%m.%Y').date() \
                if form.dob.data else None
            person_type = CHILD

        else:
            dob = None
            person_type = ADULT

        student = Person(
            last_name=last_name,
            first_name=first_name,
            patronym=patronym,
            dob=dob,
            person_type=person_type,
            status=status
        )
        message = ''

    else:
        student = None
        message = f"Человек с именем {last_name} {first_name} уже есть в системе"

    return student, message


def create_contact(form):
    telegram = form.telegram.data
    phone = form.phone.data
    other_contact = form.other_contact.data

    contact = Contact(
        telegram=telegram,
        phone=phone,
        other_contact=other_contact
    )

    return contact


def create_parent(contact_form):
    parent_last_name = contact_form.parent_last_name.data
    parent_first_name = contact_form.parent_first_name.data
    parent_patronym = contact_form.parent_patronym.data
    same_person = Person.query.filter_by(last_name=parent_last_name, first_name=parent_first_name).all()

    if not same_person:
        parent = Person(
            last_name=parent_last_name,
            first_name=parent_first_name,
            patronym=parent_patronym,
            person_type=ADULT
        )
        message = ''

    else:
        parent = None
        message = f"Человек с именем {parent_last_name} {parent_first_name} уже есть в системе"

    return parent, message


def handle_contact_info(contact_form, student):
    contact_select = contact_form.contact_select.data
    relation_type = contact_form.relation.data
    student_contacts = Contact.query.filter_by(person_id=student.id).first()
    if relation_type == CHILD_SELF:
        if not student_contacts:
            contact = create_contact(contact_form)
            db.session.add(contact)
            db.session.flush()
            student.contacts.append(contact)

        else:
            contact = student.contacts[0]
            student.contacts[0].telegram = contact_form.telegram.data
            student.contacts[0].phone = contact_form.phone.data
            student.contacts[0].other_contact = contact_form.other_contact.data

        db.session.flush()

    else:
        if contact_select == CHOOSE:
            parent_id = int(contact_form.selected_contact.data)
            parent = Person.query.filter_by(id=parent_id).first()
            contact = Contact.query.filter_by(person_id=parent_id).first()

        else:
            parent, message = create_parent(contact_form)
            if parent:
                contact = create_contact(contact_form)
                db.session.add(parent)
                db.session.add(contact)
                db.session.flush()

                parent.contacts.append(contact)
                parent.primary_contact = parent.id

            else:
                return message

        student.parents.append(parent)
        db.session.flush()
        assign_relation_type(contact_form, student, parent)

    if contact_form.primary_contact.data:
        student.primary_contact = contact.person_id

    return ''


def assign_relation_type(form, student, parent):
    relation_type = form.relation.data

    if relation_type == OTHER:
        relation_type = form.other_relation.data

    relation_entry = parent_child_table.update().where(
        (parent_child_table.c.parent_id == parent.id) &
        (parent_child_table.c.child_id == student.id)
    ).values(relation=relation_type)

    db.session.execute(relation_entry)


def add_child(form):
    student, message = create_student(form, 'child')
    if student:
        db.session.add(student)
        db.session.flush()

        for contact_form in form.contacts:
            message = handle_contact_info(contact_form, student)
            if message:
                return None, message

    return student, message


def add_adult(form):
    client_select = form.client_select.data
    if client_select == CHOOSE:
        client_id = int(form.selected_client.data)
        client = Person.query.filter_by(id=client_id).first()
        client.status = form.status.data
        message = ''
        db.session.flush()

    else:
        client, message = create_student(form, 'adult')
        if client:
            contact = create_contact(form)
            db.session.add(client)
            db.session.add(contact)
            db.session.flush()
            client.contacts.append(contact)
            client.primary_contact = client.id
            db.session.flush()

    return client, message


def add_new_employee(form):
    employee_select = form.client_select.data
    if employee_select == CHOOSE:
        person_id = int(form.selected_client.data)
        employee = Person.query.filter_by(id=person_id).first()

    else:
        last_name = form.last_name.data
        first_name = form.first_name.data
        same_person = Person.query.filter_by(last_name=last_name, first_name=first_name).all()
        if same_person:
            return None, f"Человек с именем {last_name} {first_name} уже есть в системе"

        else:
            patronym = form.patronym.data
            employee = Person(
                last_name=last_name,
                first_name=first_name,
                patronym=patronym,
                person_type=ADULT
            )
            contact = create_contact(form)
            db.session.add(employee)
            db.session.add(contact)
            db.session.flush()
            employee.contacts.append(contact)
            employee.primary_contact = employee.id

    roles = form.roles.data
    if roles:
        for role in roles:
            employee_role = role[0].upper() + role[1:]
            new_employee = Employee(
                person_id=employee.id,
                role=employee_role
            )
            db.session.add(new_employee)
            if role == TEACHER:
                employee.teacher = True
                subject_ids = [int(subject_id) for subject_id in form.subjects.data]
                subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
                employee.subjects_taught.extend(subjects)
                classes_ids = [int(class_id) for class_id in form.school_classes.data]
                school_classes = SchoolClass.query.filter(SchoolClass.id.in_(classes_ids)).all()
                employee.teaching_classes.extend(school_classes)
                employee.color = form.teacher_color.data

    return employee, ''


def clients_data(person_type):
    if person_type == 'child':
        all_clients = Person.query.filter(
            Person.contacts.any(~Contact.id.is_(None))
        ).order_by(Person.last_name, Person.first_name).all()
    elif person_type == 'adult':
        all_clients = Person.query.filter(Person.status.is_(None)).order_by(Person.last_name, Person.first_name).all()
    else:
        all_clients = Person.query.filter(
            ~Person.roles.any(Employee.id),
            Person.person_type == ADULT
        ).order_by(Person.last_name, Person.first_name).all()
    clients = []
    for client in all_clients:
        client_data = {
            "id": client.id,
            "last_name": client.last_name,
            "first_name": client.first_name,
            "patronym": client.patronym
        }
        if client.contacts:
            client_data["telegram"] = client.contacts[0].telegram
            client_data["phone"] = client.contacts[0].phone
            client_data["other_contact"] = client.contacts[0].other_contact
        else:
            client_data["telegram"] = ""
            client_data["phone"] = ""
            client_data["other_contact"] = ""

        clients.append(client_data)

    return clients


def format_status(student):
    if student.status == "Закрыт":
        student.status_info = f"{student.status} причина: {student.leaving_reason}"
    elif student.status == "Пауза":
        student.status_info = f"{student.status} до {student.pause_date}" if student.pause_date else student.status
    else:
        student.status_info = student.status


def format_student_info(student):
    if student.dob:
        dob = student.dob
        student.birth_date = f'{dob:%d.%m.%Y}'
        student.age = person_age(dob)
    if student.pause_until:
        student.pause_date = f'{student.pause_until:%d.%m.%y}'

    format_status(student)

    if student.balance > 0:
        student.balance_plus = round(student.balance, 1)
    elif student.balance < 0:
        student.balance_minus = round(student.balance, 1)
    if student.children.all():
        format_children(student)
    if student.roles:
        student.employee_roles = ', '.join([role.role for role in student.roles])


def format_main_contact(student):
    main_contact = Person.query.filter_by(id=student.primary_contact).first()
    if main_contact.contacts[0].telegram:
        student.contact = f"Телеграм: {main_contact.contacts[0].telegram}"
    elif main_contact.contacts[0].phone:
        student.contact = f"Тел.: {main_contact.contacts[0].phone}"
    elif main_contact.contacts[0].other_contact:
        student.contact = main_contact.contacts[0].other_contact

    if main_contact.id != student.id:
        contact_type = db.session.query(parent_child_table.c.relation).filter(
            parent_child_table.c.parent_id == main_contact.id,
            parent_child_table.c.child_id == student.id).scalar()
        student.contact_name = f"{contact_type} - {main_contact.last_name} {main_contact.first_name}"


def format_all_contacts(student):
    contacts = []

    if student.contacts:
        if student.primary_contact == student.id:
            student.main_contact = student
            student.main_contact.type = CHILD_SELF
        else:
            student.type = CHILD_SELF
            contacts.append(student)

    for parent in student.parents.all():
        parent.type = db.session.query(parent_child_table.c.relation).filter(
            parent_child_table.c.parent_id == parent.id,
            parent_child_table.c.child_id == student.id).scalar()
        if student.primary_contact == parent.id:
            student.main_contact = parent
        else:
            contacts.append(parent)

    student.additional_contacts = contacts


def format_children(person):
    children = []
    for child in person.children.all():
        child_age = person_age(child.dob) if child.dob else None
        child_info = f'{child.last_name} {child.first_name} ({child_age})' \
            if child_age else f'{child.last_name} {child.first_name}'
        children.append((child.id, child_info))
    person.children_info = children


def after_school_subject():
    after_school = Subject.query.filter(Subject.subject_type.has(SubjectType.name == 'after_school')).first()
    return after_school


def format_subjects_and_subscriptions(student):
    check_subscription(student, 0, 0)
    subscriptions = []
    subscriptions_list = []
    subscriptions_set = set()
    after_school_list = []

    school = [student.school_class.school_name] if student.school_class else []

    for subscription in student.subscriptions:
        subject = subscription.subject
        is_after_school = subject == after_school_subject()
        is_active = subscription.active
        subscription_dict = {}

        if is_active and not is_after_school:
            subscription_dict['subscription_id'] = subscription.id
            subscription_dict['subject_name'] = subject.name
            subscription_dict['lessons_left'] = subscription.lessons_left
            subscription_dict['purchase_date'] = f'{subscription.purchase_date:%d.%m.%Y}'
            subscription_dict['end_date'] = f'{subscription.end_date:%d.%m.%Y}'
            subscription_dict['full_subscription'] = True \
                if subscription.lessons_left == subscription.subscription_type.lessons else False
            subscriptions.append(subscription_dict)
            subscriptions_list.append(f'{subject.name}({subscription.lessons_left})')
            subscriptions_set.add(subject.id)

        elif is_after_school:
            if is_active:
                subscription_dict['subscription_id'] = subscription.id
                subscription_dict['purchase_date'] = subscription.purchase_date
                subscription_dict['shift'] = subscription.shift
                if subscription.period == "month":
                    subscription_dict['period'] = "месяц"
                    subscription_dict['validity'] = MONTHS[subscription.purchase_date.month - 1]
                elif subscription.period == "week":
                    subscription_dict['period'] = "неделя"
                    subscription_dict['validity'] = f'{subscription.purchase_date:%d.%m}-' \
                                                    f'{subscription.end_date:%d.%m.%y}'
                after_school_list.append(subscription_dict)
                subscriptions_list.insert(0, f'{subject.name} ({subscription_dict["validity"]})')
                subscriptions_set.add(subject.id)
            else:
                day_delta = (subscription.purchase_date - get_today_date()).days
                if (-30 <= day_delta <= 30) and subscription.period != "month":
                    subscription_dict['subscription_id'] = subscription.id
                    subscription_dict['purchase_date'] = subscription.purchase_date
                    subscription_dict['shift'] = subscription.shift
                    subscription_dict['period'] = "день" if subscription.period == "day" else "неделя" \
                        if subscription.period == "week" else subscription.period
                    subscription_dict['validity'] = f'{subscription.purchase_date:%d.%m}-' \
                                                    f'{subscription.end_date:%d.%m.%y}' \
                        if subscription.period == "week" else f'{subscription.purchase_date:%d.%m}'

                    after_school_list.append(subscription_dict)
                    subscriptions_set.add(subject.id)

    all_subjects_list = sorted(
        [(subject.name, subject.id) for subject in student.subjects if subject.subject_type.name != "school"])
    extra_subjects = sorted([subject.name for subject in student.subjects if
                             subject.id not in subscriptions_set and subject.subject_type.name != "school"])

    student.extra_subjects = extra_subjects
    student.subjects_info = ', '.join(school + subscriptions_list + extra_subjects)
    student.subscriptions_info = subscriptions
    student.after_school_info = after_school_list
    student.all_subjects = ([(school[0], 0)] if school else []) + all_subjects_list


def basic_student_info(student):
    format_student_info(student)
    format_main_contact(student)
    format_subjects_and_subscriptions(student)


def extensive_student_info(student):
    format_student_info(student)
    format_all_contacts(student)
    format_subjects_and_subscriptions(student)


def subscription_subjects_data():
    filtered_subjects = Subject.query.filter(
        Subject.id != after_school_subject().id,
        Subject.subscription_types.any(SubscriptionType.id.isnot(None))
    ).order_by(Subject.name).all()

    subscription_subjects = []
    for subject in filtered_subjects:
        subject_data = {
            "id": subject.id,
            "name": subject.name,
            "price_info": {subscription_type.id: f"{subscription_type.price:.0f} Лари"
                           for subscription_type in subject.subscription_types},
            "subscription_types_info": {
                subscription_type.id: f"{conjugate_lessons(subscription_type.lessons)} " +
                                      f"на {conjugate_days(subscription_type.duration)} " +
                                      f"({subscription_type.price:.0f} Лари)"

                for subscription_type in subject.subscription_types
            }
        }
        subscription_subjects.append(subject_data)

    return subscription_subjects


def potential_client_subjects():
    school_classes = SchoolClass.query.order_by(SchoolClass.school_name).all()
    potential_classes = [(school_class.id, school_class.school_name) for school_class in school_classes
                         if school_class.school_name.endswith("класс")]
    all_subjects = Subject.query.filter(
        Subject.subject_type.has(SubjectType.name.in_(["extra", "after_school", "individual"]))
    ).order_by(Subject.name).all()
    potential_subjects = [(subject.id, f"{subject.name} ({subject.subject_type.description})")
                          for subject in all_subjects]

    return {"school": potential_classes, "subjects": potential_subjects}


def subjects_data():
    all_subjects = Subject.query.filter(
        Subject.subject_type.has(SubjectType.name != 'event')
    ).order_by(Subject.name).all()
    subjects_teachers = []
    for subject in all_subjects:
        subject_data = {
            "id": subject.id,
            "name": subject.name,
            "description": subject.subject_type.description,
            "subject_type": subject.subject_type_id,
            "school_classes": {school_class.id: school_class.school_name
                               for school_class in sorted(subject.school_classes, key=lambda x: x.school_class)},
            "teachers": {teacher.id: f"{teacher.last_name} {teacher.first_name}"
                         for teacher in sorted(subject.teachers, key=lambda x: (x.last_name, x.first_name))}
        }
        subjects_teachers.append(subject_data)

    return subjects_teachers


def lesson_subjects_data():
    now = datetime.now(LOCAL_TZ).time()
    filtered_subjects = Subject.query.filter(
        Subject.subject_type.has(SubjectType.name == "extra")
    ).order_by(Subject.name).all()

    lesson_subjects = []
    for subject in filtered_subjects:
        future_lessons = Lesson.query.filter(
            and_(
                Lesson.subject_id == subject.id,
                or_(
                    and_(
                        Lesson.date == get_today_date(),
                        Lesson.start_time > now
                    ),
                    Lesson.date > get_today_date()
                )
            )
        ).order_by(Lesson.date, Lesson.start_time).all()
        if future_lessons:
            lessons_list = {lesson.id: f"{DAYS_OF_WEEK[lesson.date.weekday()]} " +
                                       f"{lesson.date:%d.%m} " +
                                       f"в {lesson.start_time:%H:%M}"
                            for lesson in future_lessons}
            subject_data = {
                "id": subject.id,
                "name": subject.name,
                "lessons": lessons_list
            }
            lesson_subjects.append(subject_data)
    return lesson_subjects


def student_lesson_register(form, student):
    subject_id = int(form.get('selected_subject')) if form.get('selected_subject') else None
    subject = Subject.query.filter_by(id=subject_id).first()
    lesson_id = int(form.get('lesson')) if form.get('lesson') else None
    lesson = Lesson.query.filter_by(id=lesson_id).first()
    if student and subject and lesson:
        student_lessons = Lesson.query.filter(
            Lesson.date == lesson.date,
            Lesson.start_time < lesson.end_time,
            Lesson.end_time > lesson.start_time,
            Lesson.students_registered.any(Person.id == student.id)
        ).all()
        if student_lessons:
            return None, "Клиент уже записан на занятие в это же время"

        if student not in subject.students:
            subject.students.append(student)
        if student not in lesson.students_registered:
            lesson.students_registered.append(student)

        return lesson, "Клиент записан на занятие"

    else:
        return None, "Ошибка при записи клиента"


def purchase_subscription(form):
    student_id = int(form.get('student_id'))
    student = Person.query.filter_by(id=student_id, status="Клиент").first()
    subject_id = int(form.get('selected_subject'))
    subscription_type_id = int(form.get('subscription_type'))
    subscription_type = SubscriptionType.query.filter_by(id=subscription_type_id).first()
    purchase_date = datetime.strptime(form.get('purchase_date'), '%d.%m.%Y').date()
    end_date = purchase_date + timedelta(subscription_type.duration)
    operation_type = form.get('operation_type')

    new_subscription = Subscription(
        subject_id=subject_id,
        student_id=student.id,
        subscription_type_id=subscription_type_id,
        lessons_left=subscription_type.lessons,
        purchase_date=purchase_date,
        end_date=end_date
    )
    return new_subscription, subscription_type.price, operation_type


def handle_student_edit(form, student, edit_type, user):
    if edit_type == 'edit_student':
        last_name = form.last_name.data
        first_name = form.first_name.data
        same_person = Person.query.filter(
            Person.id != student.id,
            Person.last_name == last_name,
            Person.first_name == first_name
        ).all()
        if same_person:
            return f"Человек с именем {last_name} {first_name} уже есть в системе"

        else:
            student.last_name = last_name
            student.first_name = first_name
            student.patronym = form.patronym.data
            student.dob = datetime.strptime(form.dob.data, '%d.%m.%Y').date() \
                if form.dob.data else None
            status = form.status.data
            if student.status != status and user.rights == 'admin':
                student.status = status
                if student.status == "Закрыт":
                    student.subjects = []
                    student.subscriptions = []
                student.pause_date = datetime.strptime(form.pause_until.data, '%d.%m.%Y').date() \
                    if form.pause_until.data and student.status == "Пауза" else None

                student.leaving_reason = form.leaving_reason.data if student.status == "Закрыт" else ''
            db.session.flush()

        return ''

    if edit_type == 'edit_main_contact':
        main_contact = student.main_contact
        if main_contact.id != student.id:
            contact_last_name = form.last_name.data
            contact_first_name = form.first_name.data
            same_person = Person.query.filter(
                Person.id != main_contact.id,
                Person.last_name == contact_last_name,
                Person.first_name == contact_first_name
            ).all()
            if same_person:
                return f"Человек с именем {contact_last_name} {contact_first_name} уже есть в системе"

            main_contact.last_name = contact_last_name
            main_contact.first_name = contact_first_name
            main_contact.patronym = form.patronym.data

        main_contact.contacts[0].telegram = form.telegram.data
        main_contact.contacts[0].phone = form.phone.data
        main_contact.contacts[0].other_contact = form.other_contact.data

        db.session.flush()
        return ''

    if edit_type.startswith('edit_contact_'):
        ind = int(edit_type[len('edit_contact_'):])
        contact = student.additional_contacts[ind - 1]
        if contact.id != student.id:
            contact_last_name = form.last_name.data
            contact_first_name = form.first_name.data
            same_person = Person.query.filter(
                Person.id != contact.id,
                Person.last_name == contact_last_name,
                Person.first_name == contact_first_name
            ).all()
            if same_person:
                return f"Человек с именем {contact_last_name} {contact_first_name} уже есть в системе"

            contact.last_name = contact_last_name
            contact.first_name = contact_first_name
            contact.patronym = form.patronym.data

        contact.contacts[0].telegram = form.telegram.data
        contact.contacts[0].phone = form.phone.data
        contact.contacts[0].other_conta = form.other_contact.data

        if form.primary_contact.data:
            student.primary_contact = contact.id
        db.session.flush()
        return ''

    if edit_type == 'new_contact':
        message = handle_contact_info(form, student)
        return message

    if edit_type == 'del_subject':
        del_subject_id = int(form.get('del_subject_btn'))
        if student.status == "Лид" and del_subject_id == 0:
            student.school_class_id = None
            db.session.flush()
        else:
            del_subject = Subject.query.filter_by(id=del_subject_id).first()
            if del_subject in student.subjects:
                student.subjects.remove(del_subject)
                db.session.flush()
        return ''

    if edit_type == 'subscription':
        for subscription_form in form.subscriptions:
            subscription_id = int(subscription_form.subscription_id.data)
            subscription = Subscription.query.filter_by(id=subscription_id).first()
            lessons_left = subscription_form.lessons.data
            purchase_date = datetime.strptime(subscription_form.purchase_date.data, '%d.%m.%Y').date()
            end_date = datetime.strptime(subscription_form.end_date.data, '%d.%m.%Y').date()
            if subscription.lessons_left != lessons_left or subscription.purchase_date != purchase_date \
                    or subscription.end_date != end_date:
                subscription.lessons_left = lessons_left
                subscription.purchase_date = purchase_date
                subscription.end_date = end_date
                description = f"Изменение абонемента {subscription.subject.name} клиента " \
                              f"{student.last_name} {student.first_name}"
                user_action(user, description)
        db.session.flush()
        return

    if edit_type == 'del_after_school':
        del_after_school_id = int(form.get('del_after_school'))
        del_after_school = Subscription.query.filter_by(id=del_after_school_id).first()
        if del_after_school:
            price = del_after_school.subscription_type.price
            if del_after_school.period not in ["month", "day", "week"]:
                hours = int(del_after_school.period.split()[0])
                price *= hours
            record = Finance.query.filter(
                Finance.person_id == del_after_school.student_id,
                Finance.service == "after_school",
                Finance.service_id == del_after_school.id
            ).first()
            if record:
                description = "Возврат за продленку"
                finance_operation(del_after_school.student, abs(record.amount), record.operation_type,
                                  description, "del_after_school", None, balance=record.student_balance)
                record.service_id = None
            else:
                description = "Возврат за продленку"
                finance_operation(del_after_school.student, price, 'cash', description,
                                  "del_after_school", del_after_school.id)
            db.session.delete(del_after_school)
            db.session.flush()

        return


def handle_employee_edit(form, employee):
    last_name = form.get('last_name')
    first_name = form.get('first_name')
    same_person = Person.query.filter(
        Person.id != employee.id,
        Person.last_name == last_name,
        Person.first_name == first_name
    ).all()
    if same_person:
        return f"Человек с именем {last_name} {first_name} уже есть в системе"

    employee.last_name = last_name
    employee.first_name = first_name
    employee.patronym = form.get('patronym')
    employee.contacts[0].telegram = form.get('telegram')
    employee.contacts[0].phone = form.get('phone')
    employee.contacts[0].other_contact = form.get('other_contact')

    for role in employee.roles:
        if not form.get(f'role_{role.id}'):
            db.session.delete(role)
            if role.role == TEACHER:
                employee.teacher = False
                employee.subjects_taught = []
                employee.teaching_classes = []
        else:
            role.role = form.get(f'role_{role.id}')
    db.session.flush()

    if employee.teacher:
        subjects_to_remove = []
        for subject in employee.subjects_taught:
            if not form.get(f'subject_{subject.id}'):
                subjects_to_remove.append(subject)

        [employee.subjects_taught.remove(subj) for subj in subjects_to_remove]

        new_subject_ids = form.getlist('new_subjects')
        if new_subject_ids:
            new_subjects = Subject.query.filter(
                Subject.id.in_([int(subject_id) for subject_id in new_subject_ids])
            ).all()
            employee.subjects_taught.extend(new_subjects)

        classes_to_remove = []
        for school_class in employee.teaching_classes:
            if not form.get(f'school_class_{school_class.id}'):
                classes_to_remove.append(school_class)

        [employee.teaching_classes.remove(sc_cl) for sc_cl in classes_to_remove]

        new_classes_ids = form.getlist('new_classes')
        if new_classes_ids:
            new_classes = SchoolClass.query.filter(
                SchoolClass.id.in_([int(class_id) for class_id in new_classes_ids])
            ).all()
            employee.teaching_classes.extend(new_classes)

        employee.color = form.get('new_teacher_color')
        db.session.flush()

    new_roles = form.getlist('new_roles')
    if new_roles:
        for new_role in new_roles:
            new_employee_role = Employee(
                person_id=employee.id,
                role=new_role
            )
            db.session.add(new_employee_role)

            if new_role == TEACHER:
                employee.teacher = True
                subject_ids = form.getlist('subjects')
                if subject_ids:
                    teacher_subjects = Subject.query.filter(
                        Subject.id.in_([int(subject_id) for subject_id in subject_ids])
                    ).all()
                    employee.subjects_taught.extend(teacher_subjects)
                classes_ids = form.getlist('classes')
                if classes_ids:
                    teacher_classes = SchoolClass.query.filter(
                        SchoolClass.id.in_([int(class_id) for class_id in classes_ids])
                    ).all()
                    employee.teaching_classes.extend(teacher_classes)
                employee.color = form.get('teacher_color')

        db.session.flush()

    return ''


def format_employee(employee):
    if employee.contacts[0].telegram:
        employee.contact = f"Телеграм: {employee.contacts[0].telegram}"
    elif employee.contacts[0].phone:
        employee.contact = f"Тел.: {employee.contacts[0].phone}"
    elif employee.contacts[0].other_contact:
        employee.contact = employee.contacts[0].other_contact
    if employee.subjects_taught.all():
        school_classes = set()
        school_subjects = []
        filtered_subjects = []
        for subject in employee.subjects_taught:
            distinct_classes = db.session.query(SchoolClass) \
                .join(Lesson.school_classes) \
                .filter(Lesson.teacher_id == employee.id,
                        Lesson.subject_id == subject.id,
                        SchoolClass.school_class <= 4).all()

            if distinct_classes:
                school_classes.update([sc_cl.school_name for sc_cl in distinct_classes])
                school_subjects.append(subject.name)
            else:
                filtered_subjects.append(subject.name)
        all_subjects = (list(school_classes) if len(school_subjects) > 2 else school_subjects) + filtered_subjects
        employee.all_subjects = ', '.join(all_subjects)

    if employee.children.all():
        format_children(employee)
    if employee.status:
        format_status(employee)


def add_new_subject(form, subject_type):
    if subject_type == "extra_school":
        name = form.subject_name.data
        subject_name = name[0].upper() + name[1:]
        short_name = form.subject_short_name.data
        subject_short_name = short_name[0].upper() + short_name[1:]
        teacher_ids = [int(teacher) for teacher in form.teachers.data]
        teachers = Person.query.filter(Person.id.in_(teacher_ids), Person.teacher).all()
        subject_type_id = int(form.subject_type.data)
        one_time_price = float(form.subject_price.data)
        if not form.no_subject_school_price.data:
            subject_school_price = float(form.subject_school_price.data)
            school_price = subject_school_price if int(subject_school_price) > 0 else None
        else:
            school_price = None
        if not form.no_subscription.data:
            subscription_type_ids = [int(st) for st in form.subscription_types.data]
            subscription_types = SubscriptionType.query.filter(SubscriptionType.id.in_(subscription_type_ids)).all()
        else:
            subscription_types = []

        new_subject = Subject(
            name=subject_name,
            short_name=subject_short_name,
            subject_type_id=subject_type_id,
            one_time_price=one_time_price,
            school_price=school_price,
        )
        new_subject.subscription_types.extend(subscription_types)
        new_subject.teachers.extend(teachers)

    else:
        name = form.get("subject_name")
        subject_name = name[0].upper() + name[1:]
        short_name = form.get("subject_short_name")
        subject_short_name = short_name[0].upper() + short_name[1:]
        teacher_ids = [int(teacher) for teacher in form.getlist("teachers") if form.getlist("teachers")]
        teachers = Person.query.filter(Person.id.in_(teacher_ids), Person.teacher).all()
        school_type = SubjectType.query.filter_by(name="school").first()
        classes = [int(school_class) for school_class in form.getlist("school_classes")]
        school_classes = SchoolClass.query.filter(SchoolClass.id.in_(classes)).all()

        new_subject = Subject(
            name=subject_name,
            short_name=subject_short_name,
            subject_type_id=school_type.id
        )
        new_subject.school_classes.extend(school_classes)
        new_subject.teachers.extend(teachers)
        classes_students = Person.query.filter(Person.school_class_id.in_(classes)).all()
        new_subject.students.extend(classes_students)
        for teacher in teachers:
            [teacher.teaching_classes.append(sc_cl) for sc_cl in school_classes
             if sc_cl not in teacher.teaching_classes]

    same_subject = Subject.query.filter_by(
        name=new_subject.name,
        subject_type_id=new_subject.subject_type_id
    ).all()

    return new_subject if not same_subject else None


def format_subscription_types(subscription_types):
    types_of_subscription = []
    for subscription_type in subscription_types:
        if subscription_type.lessons:
            type_of_subscription = f"{conjugate_lessons(subscription_type.lessons)} за {subscription_type.price:.0f} " \
                                   f"({subscription_type.duration} дней)"
            types_of_subscription.append((subscription_type.id, type_of_subscription))

    return types_of_subscription


def handle_subject_edit(subject, form):
    subject.name = form.get('subject_name')
    subject.short_name = form.get('subject_short_name')
    subject.description = form.get('description')
    if not form.get('no_subject_price'):
        subject_price = float(form.get('subject_price'))
        subject.one_time_price = subject_price if int(subject_price) > 0 else None
    else:
        subject.one_time_price = None
    if not form.get('no_subject_school_price'):
        school_price = float(form.get('subject_school_price'))
        subject.school_price = school_price if int(school_price) > 0 else None
    else:
        subject.school_price = None
    subscription_type_ids = [int(st) for st in form.getlist('subscription_types')]
    subscription_types = SubscriptionType.query.filter(SubscriptionType.id.in_(subscription_type_ids)).all()
    subject.subscription_types = [subscription_type for subscription_type in subscription_types]


def check_subscription(student, lesson, subject_id):
    after_school = after_school_subject()
    cond = lesson == 0
    if cond:
        date = get_today_date()
    else:
        date = lesson.date
    if subject_id == 0:
        subscriptions = student.subscriptions
    else:
        subscriptions = Subscription.query.filter(Subscription.subject_id.in_([subject_id, after_school.id]),
                                                  Subscription.student_id == student.id).all()
    for subscription in subscriptions:
        if subscription.subject == after_school:
            cond11 = subscription.purchase_date.month == date.month
            cond12 = subscription.period == "month"
            if subscription.period == "week":
                cond2 = subscription.purchase_date <= date <= subscription.end_date
            else:
                cond2 = False
            cond3 = subscription.purchase_date > date
            subscription.active = True if (
                ((cond11 and cond12) or cond2)
                or (cond3 and cond)
            ) else False
            db.session.commit()
        else:
            cond1 = subscription.purchase_date <= date <= subscription.end_date
            cond2 = subscription.lessons_left > 0
            cond3 = subscription.purchase_date > date
            subscription.active = True if (
                (cond1 and cond2)
                or (cond3 and cond)
            ) else False
            db.session.commit()


def check_subscriptions(subscriptions):
    date = get_today_date()
    for subscription in subscriptions:
        cond1 = subscription.purchase_date <= date <= subscription.end_date
        cond2 = subscription.lessons_left > 0
        subscription.active = True if cond1 and cond2 else False
        db.session.commit()


def get_payment_options(student, subject_id, lesson):
    check_subscription(student, lesson, subject_id)
    after_school_sub = Subscription.query.filter_by(student_id=student.id,
                                                    subject_id=after_school_subject().id,
                                                    active=True).first()
    subscriptions = Subscription.query.filter(
        Subscription.student_id == student.id,
        Subscription.subject_id == subject_id,
        Subscription.lessons_left > 0,
        Subscription.active
    ).order_by(Subscription.purchase_date).all()

    payment_options = []
    if subscriptions:
        for subscription in subscriptions:
            payment_option = {
                'value': f'subscription_{subscription.id}',
                'type': 'Абонемент',
                'info': f'({subscription.lessons_left} до {subscription.end_date:%d.%m})'
            }
            payment_options.append(payment_option)
    elif after_school_sub:
        payment_option = {
            'value': 'after_school',
            'type': 'Продленка',
            'info': ''
        }
        payment_options.append(payment_option)
    else:
        payment_option = {
            'value': 'one_time',
            'type': 'Разовое',
            'info': ''
        }
        payment_options.append(payment_option)
    return payment_options


def carry_out_lesson(form, subject, lesson, user):
    if not lesson.lesson_completed:
        lesson_price = subject.one_time_price
        lesson_school_price = subject.school_price if subject.school_price else lesson_price
        if not lesson.students_registered.all():
            return 'Нет записанных клиентов', 'error'

        for student in lesson.students_registered.all():
            if form.get(f'attending_status_{student.id}') in ['attend', 'unreasonable']:
                payment_option = form.get(f'payment_option_{student.id}')
                if payment_option.startswith('subscription'):
                    subscription_id = int(payment_option[len('subscription_'):])
                    subscription = Subscription.query.filter_by(id=subscription_id).first()
                    if subscription.lessons_left > 0:
                        subscription.lessons_left -= 1
                        payment_info = ("Абонемент", subscription.lessons_left)
                    else:
                        description = f"Списание за занятие {subject.name}"
                        finance_operation(student, -lesson_price, 'balance', description, 'lesson', lesson.id)
                        payment_info = ("Разовое", int(lesson_price))
                else:
                    price = lesson_school_price if payment_option == 'after_school' else lesson_price
                    description = f"Списание за занятие {subject.name} {lesson.date:%d.%m.%y}"
                    finance_operation(student, -price, 'balance', description, 'lesson', lesson.id)
                    payment_info = ("Продленка", int(price)) if payment_option == 'after_school' \
                        else ("Разовое", int(price))
                    subscription_id = None

                attendance = StudentAttendance(
                    date=lesson.date,
                    lesson_time=lesson.start_time,
                    student_id=student.id,
                    lesson_id=lesson.id,
                    subject_id=subject.id,
                    attending_status=form.get(f'attending_status_{student.id}'),
                    payment_method=payment_info[0],
                    price_paid=payment_info[1] if payment_info[0] in ["Продленка", "Разовое"] else None,
                    subscription_lessons=payment_info[1] if payment_info[0] == "Абонемент" else None,
                    subscription_id=subscription_id
                )
                db.session.add(attendance)
                db.session.flush()

        lesson.lesson_completed = True
        user_action(user, f"Проведение занятия {subject.name} {lesson.date:%d.%m.%Y}")
        return 'Занятие проведено', 'success'

    else:
        return 'Занятие уже проведено', 'error'


def undo_lesson(subject, lesson):
    if lesson.lesson_completed:
        lesson_price = subject.one_time_price
        lesson_school_price = subject.school_price if subject.school_price else lesson_price
        for student in lesson.students:
            attendance = StudentAttendance.query.filter_by(
                student_id=student.id,
                lesson_id=lesson.id
            ).first()

            if attendance:
                if attendance.attending_status in ['attend', 'unreasonable']:
                    payment_option = attendance.payment_method
                    if payment_option == 'Абонемент':
                        subscription_id = int(attendance.subscription_id) if attendance.subscription_id else None
                        if not subscription_id:
                            subscription = Subscription.query.filter(
                                Subscription.student_id == student.id,
                                Subscription.subject_id == subject.id,
                                Subscription.purchase_date <= lesson.date,
                                Subscription.end_date >= lesson.date
                            ).order_by(Subscription.purchase_date.desc()).first()
                        else:
                            subscription = Subscription.query.filter_by(id=subscription_id).first()
                        if subscription:
                            subscription.lessons_left += 1

                    else:
                        price = attendance.price_paid if attendance.price_paid else lesson_school_price \
                            if payment_option == 'Продленка' else lesson_price
                        record = Finance.query.filter(
                            Finance.person_id == student.id,
                            Finance.service == "lesson",
                            Finance.service_id == lesson.id
                        ).first()
                        description = f"Возврат за занятие {subject.name} {lesson.date:%d.%m.%y}"
                        if record:
                            record.service_id = None
                        finance_operation(student, price, 'balance', description, 'del_lesson', None)

                        db.session.flush()
                db.session.delete(attendance)

        lesson.lesson_completed = False
        return 'Проведение занятия отменено', 'success'

    else:
        return 'Занятие еще не проведено', 'error'


def handle_lesson(form, subject, lesson, user):
    if 'del_client_btn' in form:
        del_client_id = int(form.get('del_client_btn'))
        del_client = Person.query.filter_by(id=del_client_id).first()
        attendance = StudentAttendance.query.filter_by(student_id=del_client_id, lesson_id=lesson.id).first()
        if del_client in subject.students:
            subject.students.remove(del_client)
        if del_client in lesson.students_registered:
            lesson.students_registered.remove(del_client)
        if attendance:
            db.session.delete(attendance)
        description = f"Удаление клиента {del_client.last_name} {del_client.first_name} c занятия {subject.name}"
        user_action(user, description)
        db.session.flush()
        return 'Клиент удален', 'success'

    if 'add_client_btn' in form:
        new_client_id = int(form.get('added_client_id')) if form.get('added_client_id') else None
        if new_client_id:
            new_client = Person.query.filter_by(id=new_client_id).first()
            subject.students.append(new_client)
            description = f"Добавление клиента {new_client.last_name} {new_client.first_name} " \
                          f"на занятие {subject.name}"
            user_action(user, description)
            db.session.flush()
            return 'Клиент добавлен', 'success'
        else:
            return 'Клиент не выбран', 'error'

    if 'registered_btn' in form:
        messages = []
        for student in lesson.students:
            if (
                    student not in lesson.students_registered
                    and form.get(f'registered_{student.id}') == 'on'
            ):
                student_lessons = Lesson.query.filter(
                    Lesson.date == lesson.date,
                    Lesson.start_time < lesson.end_time,
                    Lesson.end_time > lesson.start_time,
                    Lesson.students_registered.any(Person.id == student.id)
                ).all()
                if not student_lessons:
                    lesson.students_registered.append(student)
                    description = f"Запись клиента {student.last_name} {student.first_name} " \
                                  f"на занятие {subject.name} {lesson.date:%d.%m.%Y}"
                    user_action(user, description)
                else:
                    messages.append(f"Клиент {student.last_name} {student.first_name}" +
                                    f" уже записан на занятие в это же время")
            elif (
                    student in lesson.students_registered
                    and not form.get(f'registered_{student.id}')
            ):
                lesson.students_registered.remove(student)
                description = f"Отмена записи клиента {student.last_name} {student.first_name} " \
                              f"на занятие {subject.name} {lesson.date:%d.%m.%Y}"
                user_action(user, description)

        if messages:
            message_text = '. '.join(messages)
            return message_text, 'error'

    if 'attended_btn' in form and user.rights in ["admin", "user"]:
        message = carry_out_lesson(form, subject, lesson, user)
        return message

    if 'change_btn' in form and user.rights in ["admin", "user"]:
        message = undo_lesson(subject, lesson)
        user_action(user, f"Отмена проведения занятия {subject.name} {lesson.date:%d.%m.%Y}")
        return message


def get_lesson_students(lesson):
    if lesson.lesson_completed:
        lesson_students = Person.query.filter(
            Person.lessons_attended.any(StudentAttendance.lesson_id == lesson.id)
        ).order_by(Person.last_name, Person.first_name).all()
    else:
        lesson_students = Person.query.filter(
            Person.status == "Клиент",
            or_(
                Person.lessons_registered.any(Lesson.id == lesson.id),
                Person.subjects.any(Subject.id == lesson.subject_id)
            )
        ).order_by(Person.last_name, Person.first_name).all()

    if lesson_students:
        for student in lesson_students:
            student.payment_options = get_payment_options(student, lesson.subject_id, lesson)
            if lesson.lesson_completed:
                attendance = StudentAttendance.query.filter_by(
                    student_id=student.id,
                    lesson_id=lesson.id
                ).first()
                student.attended = attendance.attending_status if attendance else 'not_attend'
                if attendance.payment_method:
                    student_payment_method = attendance.payment_method
                    student_subscription_info = f"({attendance.subscription_lessons})" \
                        if attendance.subscription_lessons is not None else ''
                    student_payment = student_payment_method + student_subscription_info
                else:
                    student_payment = "?"
                student.payment = student_payment

    return lesson_students


def prev_next_lessons(lesson):
    query = Lesson.query.filter(Lesson.subject_id == lesson.subject_id)
    previous_lesson = query.filter(
        or_(
            and_(
                Lesson.date == lesson.date,
                Lesson.start_time < lesson.start_time
            ),
            Lesson.date < lesson.date
        )
    ).order_by(Lesson.date.desc(), Lesson.start_time.desc()).first()
    next_lesson = query.filter(
        or_(
            and_(
                Lesson.date == lesson.date,
                Lesson.start_time > lesson.start_time
            ),
            Lesson.date > lesson.date
        )
    ).order_by(Lesson.date, Lesson.start_time).first()
    prev_id = previous_lesson.id if previous_lesson else None
    next_id = next_lesson.id if next_lesson else None
    return prev_id, next_id


def show_lesson(lesson_str):
    lesson_id_split = lesson_str.split('-')
    if len(lesson_id_split) != 2:
        subject_lesson = None
        subject_id = None
    else:
        if lesson_id_split[0] == '0':
            subject_id = int(lesson_id_split[1]) if lesson_id_split[1].isdigit() else None
            last_lesson = Lesson.query.filter(
                Lesson.subject_id == subject_id,
                ~Lesson.lesson_type.has(SubjectType.name.in_(["school", "after_school", "event"]))
            ).order_by(Lesson.date.desc(), Lesson.start_time.desc()).first()
            coming_lesson = Lesson.query.filter(
                Lesson.date >= get_today_date(),
                Lesson.subject_id == subject_id,
                ~Lesson.lesson_type.has(SubjectType.name.in_(["school", "after_school", "event"]))
            ).order_by(Lesson.date, Lesson.start_time).first()
            subject_lesson = coming_lesson if coming_lesson else last_lesson if last_lesson else None
        else:
            lesson_id = int(lesson_id_split[1]) if lesson_id_split[1].isdigit() else None
            subject_lesson = Lesson.query.filter(
                Lesson.id == lesson_id,
                ~Lesson.lesson_type.has(SubjectType.name.in_(["school", "after_school", "event"]))
            ).first()
            subject_id = subject_lesson.subject_id if subject_lesson else None

    if subject_lesson:
        subject_lesson.students = get_lesson_students(subject_lesson)
        subject_lesson.prev, subject_lesson.next = prev_next_lessons(subject_lesson)
        lesson_subject = subject_lesson.subject

    else:
        lesson_subject = Subject.query.filter_by(id=subject_id).first()

    return subject_lesson, lesson_subject


def format_school_classes_names(school_classes):
    num_classes = [cl.school_name[:2].strip() for cl in school_classes
                   if (cl.school_name[0].isdigit() or cl.school_name[:2].isdigit())]
    classes = [cl.school_name for cl in school_classes
               if not (cl.school_name[0].isdigit() or cl.school_name[:2].isdigit())]
    formatted_school_classes = (
        f"{'-'.join(num_classes)} класс, {', '.join(classes)}"
        if (classes and num_classes)
        else f"{'-'.join(num_classes)} класс"
        if num_classes
        else f"{', '.join(classes)}"
    )
    return formatted_school_classes


def create_lesson_dict(lesson, timetable_type):
    start_time = (lesson.start_time.hour - 9) * 60 + lesson.start_time.minute
    end_time = (lesson.end_time.hour - 9) * 60 + lesson.end_time.minute

    lesson_dict = {
        'id': lesson.id,
        'time': f'{lesson.start_time:%H:%M} - {lesson.end_time:%H:%M}',
        'start_time': start_time,
        'end_time': end_time,
        'teacher': lesson.teacher.first_name if lesson.teacher else '-',
        'completed': lesson.lesson_completed,
        'month_index': calc_month_index(lesson.date),
        'subject_full': lesson.subject.name
    }
    if timetable_type in ['general', 'teacher']:
        if lesson.lesson_type.name == 'school':
            lesson_type = format_school_classes_names(lesson.school_classes)
        elif lesson.lesson_type.name == 'individual':
            lesson_type = 'индив'
        else:
            lesson_type = ''
        if lesson.lesson_type.name == 'event':
            lesson_dict['teacher'] = ''
            lesson_dict['color'] = '#ED42C2'
        else:
            lesson_dict['color'] = lesson.teacher.color
        lesson_dict['lesson_type'] = lesson_type
        lesson_dict['subject'] = lesson.subject_names if timetable_type == 'general' else lesson.subject.name
        lesson_dict['lesson_type_name'] = lesson.lesson_type.name

    else:
        lesson_dict['room'] = lesson.room.name
        lesson_dict['room_color'] = lesson.room.color
        lesson_dict['subject'] = lesson.subject.short_name
        lesson_dict['split'] = lesson.split_classes

    return lesson_dict


def day_lessons_list(day_room_lessons):
    lessons_for_day = []
    current_lesson = day_room_lessons[0]
    current_lesson.subject_names = [current_lesson.subject.short_name]

    for next_lesson in day_room_lessons[1:]:
        if (
                current_lesson.teacher_id == next_lesson.teacher_id
                and current_lesson.school_classes.all() == next_lesson.school_classes.all()
                and current_lesson.lesson_type.name == 'school'
                and current_lesson.lesson_type.name == 'school'
        ):
            if current_lesson.subject_names[-1] != next_lesson.subject.short_name:
                current_lesson.subject_names.append(next_lesson.subject.short_name)
            current_lesson.end_time = next_lesson.end_time
        else:
            lessons_for_day.append(create_lesson_dict(current_lesson, 'general'))
            next_lesson.subject_names = [next_lesson.subject.short_name]
            current_lesson = next_lesson

    lessons_for_day.append(create_lesson_dict(current_lesson, 'general'))

    return lessons_for_day


def get_date(day_of_week, week=0):
    today = get_today_date()
    day_of_week_date = today - timedelta(days=today.weekday()) + timedelta(days=day_of_week) + week * timedelta(weeks=1)
    return day_of_week_date


def get_weekday_date(day_of_week, date=None):
    if date is None:
        date = get_today_date()
    date_of_week_day = date - timedelta(days=date.weekday()) + timedelta(days=day_of_week)
    return date_of_week_day


def get_week_dates():
    week_dates = [get_date(0), get_date(0, 1), get_date(0, -1)]
    return week_dates


def week_lessons_dict(week, rooms, lessons_type='general'):
    week_lessons = {}
    used_rooms = set()
    week_dates = []
    week_start_hour = 22
    week_end_hour = 8
    for day, weekday in enumerate(DAYS_OF_WEEK):
        lessons_date = get_date(day, week)
        lessons_date_str = f'{lessons_date:%d.%m}'
        week_dates.append(lessons_date_str)
        day_lessons = {}
        for room in rooms:
            if lessons_type == 'general':
                lessons_filtered = Lesson.query.filter_by(
                    date=lessons_date, room_id=room.id
                ).order_by(Lesson.start_time).all()
            elif lessons_type.startswith('teacher'):
                teacher_id = int(lessons_type[len('teacher_'):])
                lessons_filtered = Lesson.query.filter(
                    Lesson.date == lessons_date,
                    Lesson.room_id == room.id,
                    Lesson.teacher_id == teacher_id
                ).order_by(Lesson.start_time).all()
            elif lessons_type == 'extra':
                lessons_filtered = Lesson.query.filter(
                    Lesson.date == lessons_date,
                    Lesson.room_id == room.id,
                    Lesson.lesson_type.has(SubjectType.name == 'extra')
                ).order_by(Lesson.start_time).all()
            else:
                lessons_filtered = Lesson.query.filter(
                    Lesson.date == lessons_date,
                    Lesson.room_id == room.id,
                    Lesson.lesson_type.has(SubjectType.name == 'individual')
                ).order_by(Lesson.start_time).all()
            if lessons_filtered:
                start_hour = lessons_filtered[0].start_time.hour
                week_start_hour = start_hour if start_hour < week_start_hour else week_start_hour
                end_hour = lessons_filtered[-1].end_time.hour + 1 * bool(lessons_filtered[-1].end_time.minute)
                week_end_hour = end_hour if end_hour > week_end_hour else week_end_hour
                used_rooms.add(room.name)
                if lessons_type.startswith('teacher'):
                    lessons_for_day = []
                    for lesson in lessons_filtered:
                        lessons_for_day.append(create_lesson_dict(lesson, 'teacher'))
                    day_lessons[room.name] = lessons_for_day
                else:
                    day_lessons[room.name] = day_lessons_list(lessons_filtered) if lessons_filtered else []
        if day_lessons:
            week_lessons[weekday] = day_lessons
    if not week_lessons.get(DAYS_OF_WEEK[-1]):
        week_dates.pop(-1)
        if not week_lessons.get(DAYS_OF_WEEK[-2]):
            week_dates.pop(-1)
    return week_lessons, week_dates, used_rooms, (week_start_hour, week_end_hour)


def day_school_lessons_dict(day, week, school_classes):
    lesson_date = get_date(day - 1, week)
    lesson_date_str = f'{lesson_date:%d.%m}'
    day_start_hour = 22
    day_end_hour = 8

    class_lessons = {}
    for school_class in school_classes:
        lessons_filtered = Lesson.query.filter(
            Lesson.lesson_type.has(SubjectType.name == "school"),
            Lesson.date == lesson_date,
            Lesson.school_classes.any(SchoolClass.id == school_class.id)
        ).order_by(Lesson.start_time).all()
        if lessons_filtered:
            start_hour = lessons_filtered[0].start_time.hour
            day_start_hour = start_hour if start_hour < day_start_hour else day_start_hour
            end_hour = lessons_filtered[-1].end_time.hour + 1 * bool(lessons_filtered[-1].end_time.minute)
            day_end_hour = end_hour if end_hour > day_end_hour else day_end_hour

            class_lessons[school_class.school_name] = [create_lesson_dict(school_lesson, "school")
                                                       for school_lesson in lessons_filtered if lessons_filtered]

    return class_lessons, lesson_date_str, (day_start_hour, day_end_hour)


def check_conflicting_lessons(date, start_time, end_time, classes, room, teacher, split_classes, lesson_id=None):
    filter_conditions = [
        and_(
            Lesson.start_time < end_time,
            Lesson.end_time > start_time,
            Lesson.room_id == room
        ),
        and_(
            Lesson.start_time < end_time,
            Lesson.end_time > start_time,
            Lesson.teacher_id == teacher
        )
    ]

    if not split_classes:
        class_condition = and_(
            Lesson.start_time < end_time,
            Lesson.end_time > start_time,
            Lesson.school_classes.any(SchoolClass.id.in_(classes))
        )
        filter_conditions.append(class_condition)

    conflicting_lessons = Lesson.query.filter(
        and_(
            Lesson.id != lesson_id,
            Lesson.date == date,
            or_(*filter_conditions)
        )
    ).all()

    return conflicting_lessons


def analyze_conflicts(conflicting_lessons, room_id, teacher_id, classes):
    intersection = set()

    for conflict in conflicting_lessons:
        if conflict.room_id == room_id:
            alert = 'кабинету (' + conflict.room.name + ')'
            intersection.add(alert)
        if conflict.teacher_id == teacher_id:
            alert = 'учителю (' + f'{conflict.teacher.last_name} {conflict.teacher.first_name}' + ')'
            intersection.add(alert)
        conflict_classes = set([school_class.id for school_class in conflict.school_classes])
        intersected_classes = conflict_classes.intersection(classes)
        if intersected_classes:
            inter_classes = SchoolClass.query.filter(SchoolClass.id.in_(intersected_classes)).all()
            alert = 'классам (' + ', '.join([school_class.school_name for school_class in inter_classes]) + ')'
            intersection.add(alert)

    return intersection


def create_check_lesson(lesson_date, form, i):
    start_time = datetime.strptime(form.start_time.data, '%H : %M').time()
    end_time = datetime.strptime(form.end_time.data, '%H : %M').time()
    subject_id, lesson_type_id = form.subject.data.split('-')
    room_id = int(form.room.data)
    classes = [int(school_class) for school_class in form.school_classes.data]
    split_classes = form.split_classes.data
    teacher_id = int(form.teacher.data)
    school_type = SubjectType.query.filter_by(name='school').first().id

    if end_time <= start_time:
        lesson = None
        message_text = f'Занятие {i} не добавлено в расписание. Ошибка во времени проведения'
        return lesson, message_text

    if int(lesson_type_id) == school_type and not classes:
        lesson = None
        message_text = f'Занятие {i} не добавлено в расписание. Классы не выбраны'
        return lesson, message_text

    conflicting_lessons = check_conflicting_lessons(lesson_date, start_time, end_time,
                                                    classes, room_id, teacher_id, split_classes)

    if conflicting_lessons:
        intersection = analyze_conflicts(conflicting_lessons, room_id, teacher_id, classes)

        message_text = f'Занятие {i} не добавлено в расписание. Пересечения по ' + ', '.join(intersection)
        lesson = None

    else:
        lesson = Lesson(
            date=lesson_date,
            start_time=start_time,
            end_time=end_time,
            lesson_type_id=int(lesson_type_id),
            room_id=room_id,
            subject_id=int(subject_id),
            teacher_id=teacher_id,
        )

        school_classes = SchoolClass.query.filter(SchoolClass.id.in_(classes)).all()
        if school_classes:
            lesson.school_classes.extend(school_classes)
            lesson.split_classes = split_classes
            school_students = Person.query.filter(
                Person.status == "Клиент",
                Person.school_class_id.in_(classes)
            ).all()
            for student in school_students:
                student_lessons = Lesson.query.filter(
                    Lesson.date == lesson.date,
                    Lesson.start_time < lesson.end_time,
                    Lesson.end_time > lesson.start_time,
                    Lesson.students_registered.any(Person.id == student.id)
                ).all()
                if not student_lessons:
                    lesson.students_registered.append(student)

        message_text = f'Занятие {i} добавлено в расписание'

    return lesson, message_text


def calculate_week(date):
    return int((get_weekday_date(0, date) - get_weekday_date(0, get_today_date())).days / 7)


def add_new_lessons(form):
    lesson_date = datetime.strptime(form.lesson_date.data, '%d.%m.%Y').date()
    messages = []
    new_lessons = 0
    for i, lesson_form in enumerate(form.lessons, 1):
        new_lesson, text = create_check_lesson(lesson_date, lesson_form, i)
        if new_lesson:
            new_lessons += 1
            message = (text, 'success')
            messages.append(message)
            db.session.add(new_lesson)
            db.session.flush()
        else:
            message = (text, 'error')
            messages.append(message)

    return messages, lesson_date, new_lessons


def lesson_edit(form, lesson):
    lesson_date = datetime.strptime(form.lesson_date.data, '%d.%m.%Y').date()
    start_time = datetime.strptime(form.start_time.data, '%H : %M').time()
    end_time = datetime.strptime(form.end_time.data, '%H : %M').time()
    room_id = int(form.room.data)
    before_classes = set([school_class.id for school_class in lesson.school_classes])
    classes = [int(school_class) for school_class in form.school_classes.data] \
        if lesson.lesson_type.name == 'school' else []
    split_classes = form.split_classes.data
    teacher_id = int(form.teacher.data) if form.teacher.data else None

    if end_time <= start_time:
        message = ('Ошибка во времени проведения', 'error')
        return message

    conflicting_lessons = check_conflicting_lessons(lesson_date, start_time, end_time, classes,
                                                    room_id, teacher_id, split_classes, lesson.id)

    if conflicting_lessons:
        intersection = analyze_conflicts(conflicting_lessons, room_id, teacher_id, classes)

        message = ('Занятие не изменено. Пересечения по ' + ', '.join(intersection), 'error')

    else:
        lesson.date = lesson_date
        lesson.start_time = start_time
        lesson.end_time = end_time
        lesson.room_id = room_id
        lesson.teacher_id = teacher_id
        if lesson.lesson_type.name == 'school':
            lesson.split_classes = split_classes
            classes_set = set(classes)
            if before_classes != classes_set:
                added_classes = list(classes_set.difference(before_classes))
                removed_classes = list(before_classes.difference(classes_set))
                school_classes = SchoolClass.query.filter(SchoolClass.id.in_(classes)).all()
                lesson.school_classes = [school_class for school_class in school_classes]

                school_students = Person.query.filter(
                    Person.status == "Клиент",
                    Person.school_class_id.in_(added_classes + removed_classes)
                ).all()

                for student in school_students:
                    if student.school_class_id in removed_classes and student in lesson.students_registered:
                        lesson.students_registered.remove(student)
                        attendance = StudentAttendance.query.filter_by(
                            student_id=student.id, lesson_id=lesson.id
                        ).first()
                        if attendance:
                            db.session.delete(attendance)
                        db.session.flush()

                    if student.school_class_id in added_classes and student not in lesson.students_registered:
                        student_lessons = Lesson.query.filter(
                            Lesson.date == lesson.date,
                            Lesson.start_time < lesson.end_time,
                            Lesson.end_time > lesson.start_time,
                            Lesson.students_registered.any(Person.id == student.id)
                        ).all()
                        if not student_lessons:
                            lesson.students_registered.append(student)
                        db.session.flush()

        message = ('Занятие изменено', 'success')

    return message


def filter_lessons(form):
    week = form.get('week')

    if week.lstrip('-').isdigit():
        start_date = get_date(0, int(week))
    else:
        week_date = datetime.strptime(form.get('week_specific'), '%d.%m.%Y').date()
        start_date = get_weekday_date(0, week_date)

    next_week = form.get('next_week')

    if next_week.isdigit():
        next_start_date = get_date(0, int(next_week))
    else:
        next_week_date = datetime.strptime(form.get('next_week_specific'), '%d.%m.%Y').date()
        next_start_date = get_weekday_date(0, next_week_date)
    week_diff = int((next_start_date - start_date).days / 7)
    next_week = int((next_start_date - get_weekday_date(0, get_today_date())).days / 7)

    weekday = form.get('lessons_days')
    weekdays = form.getlist('lessons_days_specific')

    subject_types = form.get('subject_types')
    school_classes = form.getlist('school_classes')
    rooms = form.get('rooms')
    teachers = form.get('teachers')
    school_type_id = SubjectType.query.filter_by(name='school').first().id

    if weekday == "all":
        end_date = get_weekday_date(6, start_date)
        query = Lesson.query.filter(Lesson.date >= start_date,
                                    Lesson.date <= end_date)
    else:
        lessons_dates = [get_weekday_date(int(day), start_date) for day in weekdays if weekdays]
        query = Lesson.query.filter(Lesson.date.in_(lessons_dates))

    if rooms != "all":
        rooms_list = [int(room) for room in form.getlist('rooms_specific') if form.getlist('rooms_specific')]
        query = query.filter(Lesson.room_id.in_(rooms_list))
    if subject_types != "all":
        subject_types_list = [int(room) for room in form.getlist('subject_types_specific')
                              if form.getlist('subject_types_specific')]
        query = query.filter(Lesson.lesson_type_id.in_(subject_types_list))
        if school_classes != 'all' and (school_type_id in subject_types_list):
            school_classes_list = [int(school_class) for school_class in form.getlist('school_classes_specific')
                                   if form.getlist('school_classes_specific')]
            query = query.filter(Lesson.school_classes.any(SchoolClass.id.in_(school_classes_list)))
    if teachers != "all":
        teachers_list = [int(teacher) for teacher in form.getlist('teachers_specific')
                         if form.getlist('teachers_specific')]
        query = query.filter(Lesson.lesson_type_id.in_(teachers_list))

    filtered_lessons = query.all()

    return filtered_lessons, week_diff, next_week


def copy_filtered_lessons(filtered_lessons, week_diff):
    copied_lessons = 0
    conflicts = 0
    for lesson in filtered_lessons:
        date = lesson.date + timedelta(weeks=week_diff)
        start_time = lesson.start_time
        end_time = lesson.end_time
        classes = [cl.id for cl in lesson.school_classes]
        room = lesson.room_id
        teacher = lesson.teacher_id
        split_classes = lesson.split_classes

        conflicting_lessons = check_conflicting_lessons(date, start_time, end_time, classes,
                                                        room, teacher, split_classes)

        if not conflicting_lessons:
            new_lesson = Lesson(
                date=date,
                start_time=start_time,
                end_time=end_time,
                room_id=room,
                subject_id=lesson.subject_id,
                teacher_id=teacher,
                lesson_type_id=lesson.lesson_type_id
            )
            db.session.add(new_lesson)
            if lesson.lesson_type.name == "school":
                new_lesson.school_classes = SchoolClass.query.filter(SchoolClass.id.in_(classes)).all()
                lesson.split_classes = split_classes
                if lesson.students_registered.all():
                    lesson_students = lesson.students_registered.all()
                else:
                    lesson_students = Person.query.filter(Person.school_class_id.in_(classes)).all()
                for student in lesson_students:
                    if student.status == "Клиент":
                        student_lessons = Lesson.query.filter(
                            Lesson.date == new_lesson.date,
                            Lesson.start_time < new_lesson.end_time,
                            Lesson.end_time > new_lesson.start_time,
                            Lesson.students_registered.any(Person.id == student.id)
                        ).all()
                        if not student_lessons:
                            new_lesson.students_registered.append(student)

            db.session.flush()
            copied_lessons += 1

        else:
            conflicts += 1

    return copied_lessons, conflicts


def format_school_class_students(school_class):
    school_class.main_teacher = Person.query.filter_by(id=school_class.main_teacher_id).first()
    class_students = Person.query.filter_by(
        school_class_id=school_class.id,
        status="Клиент"
    ).order_by(Person.last_name, Person.last_name).all()
    for student in class_students:
        format_student_info(student)
        format_main_contact(student)
    school_class.class_students = class_students


def format_school_class_subjects(school_class):
    school_class.main_teacher = Person.query.filter_by(id=school_class.main_teacher_id).first()
    school_class_subjects = Subject.query.filter(
        Subject.school_classes.any(SchoolClass.id == school_class.id)
    ).order_by(Subject.name).all()
    for school_subject in school_class_subjects:
        school_teachers = Person.query.filter(
            Person.lessons.any(
                and_(
                    Lesson.subject_id == school_subject.id,
                    Lesson.school_classes.any(SchoolClass.id == school_class.id)
                )
            ),
            Person.subjects_taught.any(Subject.id == school_subject.id)
        ).order_by(Person.last_name, Person.first_name).all()

        subject_teachers = Person.query.filter(
            Person.teaching_classes.any(SchoolClass.id == school_class.id),
            Person.subjects_taught.any(Subject.id == school_subject.id)
        ).order_by(Person.last_name, Person.first_name).all()

        school_subject.school_teachers = school_teachers if school_teachers else subject_teachers
    school_class.school_class_subjects = school_class_subjects


def prev_next_school_lessons(lesson):
    query = Lesson.query.filter(Lesson.subject_id == lesson.subject_id)
    for school_class in lesson.school_classes:
        query = query.filter(Lesson.school_classes.any(SchoolClass.id == school_class.id))
    previous_lesson = query.filter(
        or_(
            and_(
                Lesson.date == lesson.date,
                Lesson.start_time < lesson.start_time
            ),
            Lesson.date < lesson.date
        )
    ).order_by(Lesson.date.desc(), Lesson.start_time.desc()).first()
    next_lesson = query.filter(
        or_(
            and_(
                Lesson.date == lesson.date,
                Lesson.start_time > lesson.start_time
            ),
            Lesson.date > lesson.date
        )
    ).order_by(Lesson.date, Lesson.start_time).first()
    prev_id = previous_lesson.id if previous_lesson else None
    next_id = next_lesson.id if next_lesson else None
    return prev_id, next_id


def show_school_lesson(lesson_str):
    lesson_id_split = lesson_str.split('-')
    if len(lesson_id_split) != 2:
        sc_lesson = None
        subject_id = None
    else:
        if lesson_id_split[0] == '0':
            lesson_id = int(lesson_id_split[1]) if lesson_id_split[1].isdigit() else None
            sc_lesson = Lesson.query.filter(
                Lesson.id == lesson_id,
                Lesson.lesson_type.has(SubjectType.name == "school")
            ).first()
            subject_id = sc_lesson.subject_id if sc_lesson else None
        else:
            school_class_id = int(lesson_id_split[0]) if lesson_id_split[0].isdigit() else None
            subject_id = int(lesson_id_split[1]) if lesson_id_split[1].isdigit() else None
            last_lesson = Lesson.query.filter(
                Lesson.subject_id == subject_id,
                Lesson.lesson_type.has(SubjectType.name == "school"),
                Lesson.school_classes.any(SchoolClass.id == school_class_id)
            ).order_by(Lesson.date.desc(), Lesson.start_time.desc()).first()
            coming_lesson = Lesson.query.filter(
                Lesson.date >= get_today_date(),
                Lesson.subject_id == subject_id,
                Lesson.lesson_type.has(SubjectType.name == "school"),
                Lesson.school_classes.any(SchoolClass.id == school_class_id)
            ).order_by(Lesson.date, Lesson.start_time).first()
            sc_lesson = coming_lesson if coming_lesson else last_lesson if last_lesson else None

    if sc_lesson:
        classes = [cl.id for cl in sc_lesson.school_classes]
        sc_lesson.classes_ids = classes
        sc_lesson.classes = ', '.join([cl.school_name for cl in sorted(sc_lesson.school_classes,
                                                                       key=lambda x: x.school_class)])

        if sc_lesson.students_registered.all():
            lesson_students = Person.query.filter(
                Person.lessons_registered.any(Lesson.id == sc_lesson.id),
                Person.status == "Клиент"
            ).order_by(Person.last_name, Person.first_name).all()

        else:
            lesson_students = Person.query.filter(
                Person.school_class_id.in_(classes),
                Person.subjects.any(Subject.id == sc_lesson.subject_id),
                Person.status == "Клиент"
            ).order_by(Person.last_name, Person.first_name).all()

        for student in lesson_students:
            journal_record = SchoolLessonJournal.query.filter_by(
                lesson_id=sc_lesson.id, student_id=student.id
            ).first()
            student.grade = journal_record.grade if journal_record else None
            student.lesson_comment = journal_record.lesson_comment if journal_record else None
            attendance = StudentAttendance.query.filter_by(
                student_id=student.id,
                lesson_id=sc_lesson.id
            ).first()
            student.attending_status = attendance.attending_status if attendance else 'not_attend'
        sc_lesson.lesson_students = lesson_students
        sc_lesson.prev, sc_lesson.next = prev_next_school_lessons(sc_lesson)

        sc_subject = sc_lesson.subject

    else:
        sc_subject = Subject.query.filter_by(id=subject_id).first() if subject_id else None

    return sc_lesson, sc_subject


def handle_school_lesson(form, lesson, user):
    if 'del_student_btn' in form:
        del_student_id = int(form.get('del_student_btn'))
        del_student = Person.query.filter_by(id=del_student_id).first()
        attendance = StudentAttendance.query.filter_by(student_id=del_student.id, lesson_id=lesson.id).first()
        if del_student in lesson.students_registered:
            lesson.students_registered.remove(del_student)
        if attendance:
            db.session.delete(attendance)
        description = f"Удаление ученика {del_student.last_name} {del_student.first_name} " \
                      f"из списка учеников урока {lesson.subject.name} {lesson.date:%d.%m.%Y}"
        user_action(user, description)
        db.session.flush()

    if 'add_student_btn' in form:
        new_student_id = int(form.get('added_student_id')) if form.get('added_student_id') else None
        new_student = Person.query.filter_by(id=new_student_id).first()
        if new_student:
            new_student_lessons = Lesson.query.filter(
                Lesson.date == lesson.date,
                Lesson.start_time < lesson.end_time,
                Lesson.end_time > lesson.start_time,
                Lesson.students_registered.any(Person.id == new_student.id)
            ).all()
            if not new_student_lessons:
                if new_student not in lesson.students_registered.all():
                    lesson.students_registered.append(new_student)
                if lesson.subject not in new_student.subjects.all():
                    new_student.subjects.append(lesson.subject)
                description = f"Добавление ученика {new_student.last_name} {new_student.first_name} " \
                              f"в список учеников урока {lesson.subject.name} {lesson.date:%d.%m.%Y}"
                user_action(user, description)
                db.session.flush()
                return "Ученик добавлен", 'success'

            else:
                return "Ученик записан на другое занятие в это же время", 'error'

    if 'complete_btn' in form:
        if not lesson.lesson_completed:
            lesson.lesson_topic = form.get('lesson_topic')
            lesson.lesson_comment = form.get('comment')

            for student in lesson.lesson_students:
                if student not in lesson.students_registered:
                    lesson.students_registered.append(student)

                attendance = StudentAttendance.query.filter_by(
                    student_id=student.id,
                    lesson_id=lesson.id
                ).first()
                new_attending_status = "attend" if form.get(f'attended_{student.id}') else "not_attend"
                if not attendance:
                    new_attendance = StudentAttendance(
                        student_id=student.id,
                        lesson_id=lesson.id,
                        attending_status=new_attending_status
                    )
                    db.session.add(new_attendance)
                    db.session.flush()
                else:
                    attendance.attending_status = new_attending_status
                    db.session.flush()

                grade = form.get(f'grade_{student.id}')
                lesson_comment = form.get(f'lesson_comment_{student.id}')
                journal_record = SchoolLessonJournal.query.filter_by(
                    lesson_id=lesson.id,
                    student_id=student.id
                ).first()
                if grade or lesson_comment:
                    if journal_record:
                        journal_record.grade = grade
                        journal_record.lesson_comment = lesson_comment
                    else:
                        new_journal_record = SchoolLessonJournal(
                            date=lesson.date,
                            lesson_id=lesson.id,
                            student_id=student.id,
                            school_class_id=student.school_class_id,
                            subject_id=lesson.subject_id,
                            grade=grade,
                            lesson_comment=lesson_comment
                        )
                        db.session.add(new_journal_record)
                else:
                    if journal_record:
                        db.session.delete(journal_record)
                db.session.flush()
            lesson.lesson_completed = True
            user_action(user, f"Проведение урока {lesson.subject.name} {lesson.date:%d.%m.%Y} и заполнение журнала")

            return 'Журнал заполнен', 'success'

        else:
            return 'Журнал уже был заполнен', 'error'

    if 'change_btn' in form:
        lesson.lesson_completed = False
        description = f"Отмена проведения урока {lesson.subject.name} {lesson.date:%d.%m.%Y} для внесения изменений"
        user_action(user, description)

        return 'Внесение изменений', 'success'


def lesson_duration(lesson):
    return (lesson.end_time.hour * 60 + lesson.end_time.minute) - \
           (lesson.start_time.hour * 60 + lesson.start_time.minute)


def employee_record(employees, week):
    week_start = get_date(0, week)
    week_end = get_date(6, week)
    employees_list = []

    def add_employee_dict(employee_id, name, role, activity):
        employees_list.append({
            'id': employee_id,
            'name': name,
            'role': role,
            'activity': activity
        })

    for employee in employees:
        for role in employee.roles:
            if role.role != "Учитель":
                add_employee_dict(employee.id, f"{employee.last_name} {employee.first_name}", role.role, {})

        if employee.teacher:
            teacher_lessons = Lesson.query.filter(
                Lesson.date >= week_start,
                Lesson.date <= week_end,
                Lesson.teacher_id == employee.id
            ).all()
            main_teacher = SchoolClass.query.filter(
                SchoolClass.school_class < 5,
                SchoolClass.main_teacher_id == employee.id
            ).all()
            lessons_dict = {}
            if main_teacher:
                teacher_classes_ids = [sc_cl.id for sc_cl in main_teacher]
                for lesson in teacher_lessons:
                    lesson_date = f'{lesson.date:%d.%m}'
                    if len(lesson.school_classes.all()) == 1 and (lesson.school_classes[0].id in teacher_classes_ids):
                        if lessons_dict.get(lesson.school_classes[0].school_name):
                            if lessons_dict[lesson.school_classes[0].school_name].get(lesson_date):
                                lessons_dict[lesson.school_classes[0].school_name][lesson_date] += 1
                            else:
                                lessons_dict[lesson.school_classes[0].school_name][lesson_date] = 1
                        else:
                            lessons_dict[lesson.school_classes[0].school_name] = {lesson_date: 1}
                    else:
                        if lessons_dict.get(lesson.subject.name):
                            if lessons_dict[lesson.subject.name].get(lesson_date):
                                if lesson.lesson_type.name != 'school':
                                    lessons_dict[lesson.subject.name][lesson_date] += lesson_duration(lesson)
                                else:
                                    lessons_dict[lesson.subject.name][lesson_date] += 1
                            else:
                                if lesson.lesson_type.name != 'school':
                                    lessons_dict[lesson.subject.name][lesson_date] = lesson_duration(lesson)
                                else:
                                    lessons_dict[lesson.subject.name][lesson_date] = 1
                        else:
                            if lesson.lesson_type.name != 'school':
                                lessons_dict[lesson.subject.name] = {lesson_date: lesson_duration(lesson)}
                            else:
                                lessons_dict[lesson.subject.name] = {lesson_date: 1}
            else:
                for lesson in teacher_lessons:
                    lesson_date = f'{lesson.date:%d.%m}'
                    if lessons_dict.get(lesson.subject.name):
                        if lessons_dict[lesson.subject.name].get(lesson_date):
                            if lesson.lesson_type.name != 'school':
                                lessons_dict[lesson.subject.name][lesson_date] += lesson_duration(lesson)
                            else:
                                lessons_dict[lesson.subject.name][lesson_date] += 1
                        else:
                            if lesson.lesson_type.name != 'school':
                                lessons_dict[lesson.subject.name][lesson_date] = lesson_duration(lesson)
                            else:
                                lessons_dict[lesson.subject.name][lesson_date] = 1
                    else:
                        if lesson.lesson_type.name != 'school':
                            lessons_dict[lesson.subject.name] = {lesson_date: lesson_duration(lesson)}
                        else:
                            lessons_dict[lesson.subject.name] = {lesson_date: 1}
            for role, activity in lessons_dict.items():
                add_employee_dict(employee.id, f"{employee.last_name} {employee.first_name}", role, activity)

    dates = [f"{week_start + timedelta(day):%d.%m}" for day in range(7)]

    return employees_list, dates


def school_subject_record(subject_id, school_classes_ids, month_index):
    result_date = get_today_date() + relativedelta(months=month_index)
    first_date = datetime(result_date.year, result_date.month, 1).date()
    last_date = first_date + relativedelta(months=+1, days=-1)
    school_students = Person.query.filter(
        Person.school_class_id.in_(school_classes_ids)
    ).order_by(Person.last_name, Person.first_name).all()
    subject_records = SchoolLessonJournal.query.filter(
        SchoolLessonJournal.date >= first_date,
        SchoolLessonJournal.date <= last_date,
        SchoolLessonJournal.subject_id == subject_id,
        SchoolLessonJournal.school_class_id.in_(school_classes_ids),
        ~SchoolLessonJournal.final_grade
    ).all()
    dates_topic_set = set()
    record_dict = {f"{student.last_name} {student.first_name}": {} for student in school_students}

    for record in subject_records:
        date_string = f"{record.date:%d.%m}"
        topic = record.grade_type if record.grade_type else record.lesson.lesson_topic
        lesson_id = record.lesson_id if record.lesson_id else 0
        dates_topic_set.add((date_string, topic, lesson_id))
        student_name = f"{record.student.last_name} {record.student.first_name}"
        if record.student in school_students:
            record_dict[student_name][(date_string, topic, lesson_id)] = {
                "grade": record.grade,
                "comment": record.lesson_comment
            }

    subject_lessons = Lesson.query.filter(
        Lesson.date >= first_date,
        Lesson.date <= last_date,
        Lesson.subject_id == subject_id,
        Lesson.school_classes.any(SchoolClass.id.in_(school_classes_ids)),
        Lesson.lesson_completed
    ).all()
    for lesson in subject_lessons:
        date_string = f"{lesson.date:%d.%m}"
        topic = lesson.lesson_topic
        dates_topic_set.add((date_string, topic, lesson.id))

    dates_topic = sorted(list(dates_topic_set))

    school_start_year = first_date.year if 9 <= first_date.month <= 12 else first_date.year - 1
    final_grades = SchoolLessonJournal.query.filter(
        SchoolLessonJournal.date >= datetime(school_start_year, 9, 1).date(),
        SchoolLessonJournal.date <= datetime(school_start_year + 1, 7, 1).date(),
        SchoolLessonJournal.subject_id == subject_id,
        SchoolLessonJournal.school_class_id.in_(school_classes_ids),
        SchoolLessonJournal.final_grade
    ).order_by(SchoolLessonJournal.date).all()

    final_grades_list = []
    if final_grades:
        final_grades_set = set()
        for final_grade in final_grades:
            date_string = f"{final_grade.date:%d.%m}"
            topic = final_grade.grade_type
            lesson_id = 0
            final_grades_set.add((date_string, topic, lesson_id))
            student_name = f"{final_grade.student.last_name} {final_grade.student.first_name}"
            if final_grade.student in school_students:
                record_dict[student_name][(date_string, topic, lesson_id)] = {
                    "grade": final_grade.grade,
                    "comment": final_grade.lesson_comment
                }

        final_grades_list = sorted(list(final_grades_set))

    return record_dict, dates_topic, school_students, final_grades_list


def subject_record(subject_id, month_index):
    result_date = get_today_date() + relativedelta(months=month_index)
    first_date = datetime(result_date.year, result_date.month, 1).date()
    last_date = first_date + relativedelta(months=+1, days=-1)

    subject_records = StudentAttendance.query.filter(
        StudentAttendance.date >= first_date,
        StudentAttendance.date <= last_date,
        StudentAttendance.subject_id == subject_id
    ).order_by(StudentAttendance.date, StudentAttendance.lesson_time).all()

    subject_students = []
    lesson_datetimes = []
    record_dict = {}
    for record in subject_records:
        date_time_string = f"{record.date:%d.%m} {record.lesson_time:%H:%M}"
        lesson_date_time = (date_time_string, record.lesson_id)
        subject_student = (f"{record.student.last_name} {record.student.first_name}", record.student_id)
        if record.payment_method:
            student_payment_method = record.payment_method
            student_subscription_info = f"({record.subscription_lessons})" \
                if record.subscription_lessons is not None else ''
            student_payment = student_payment_method + student_subscription_info
        else:
            student_payment = "?"
        if subject_student not in subject_students:
            subject_students.append(subject_student)
            record_dict[subject_student] = {lesson_date_time: student_payment}
        else:
            record_dict[subject_student][lesson_date_time] = student_payment
        if lesson_date_time not in lesson_datetimes:
            lesson_datetimes.append(lesson_date_time)

    sorted_subject_students = sorted(subject_students, key=lambda x: x[0])
    month = MONTHS[first_date.month - 1]

    return record_dict, lesson_datetimes, sorted_subject_students, month


def add_new_grade(form, students, subject_id, grade):
    grade_type = form.get('grade_type')
    grade_date = datetime.strptime(form.get('grade_date'), '%d.%m.%Y').date()
    for student in students:
        new_grade = form.get(f'new_grade_{student.id}')
        new_comment = form.get(f'new_comment_{student.id}')
        if new_grade or new_comment:
            journal_record = SchoolLessonJournal(
                date=grade_date,
                grade_type=grade_type,
                student_id=student.id,
                school_class_id=student.school_class_id,
                subject_id=subject_id,
                grade=new_grade,
                lesson_comment=new_comment
            )
            if grade == "final":
                journal_record.final_grade = True
            db.session.add(journal_record)
            db.session.flush()

    return grade_date, grade_type


def change_grade(form, subject, classes_ids, month_index, user):
    result_date = get_today_date() + relativedelta(months=month_index)
    month, year = (result_date.month, result_date.year)
    grade_date_topic = form.get('grade_date_topic')
    if not grade_date_topic:
        return 'Оценка не выбрана', 'error'

    date, topic, lesson_id = grade_date_topic.split('-')
    if lesson_id != 'final':
        record_date = datetime(year, month, int(date.split('.')[0])).date()
        lesson_id = int(lesson_id) if lesson_id.isdigit() and lesson_id != '0' else None
        final_grade = False
    else:
        school_start_year = year if 9 <= month <= 12 else year - 1
        record_day, record_month = map(int, date.split('.'))
        record_year = school_start_year if 9 <= record_month <= 12 else school_start_year + 1
        record_date = datetime(record_year, record_month, record_day).date()
        lesson_id = None
        final_grade = True
    if form.get('change_mode') == 'delete':
        records = SchoolLessonJournal.query.filter(
            SchoolLessonJournal.date == record_date,
            SchoolLessonJournal.grade_type == topic,
            SchoolLessonJournal.subject_id == subject.id,
            SchoolLessonJournal.school_class_id.in_(classes_ids),
            SchoolLessonJournal.lesson_id == lesson_id,
            SchoolLessonJournal.final_grade == final_grade
        ).all()
        [db.session.delete(record) for record in records]
        description = f"Удаление оценок по предмету {subject.name} от {record_date:%d.%m.%Y} ({topic})"
        user_action(user, description)
        db.session.flush()
        return 'Оценки удалены', 'success'

    else:
        student_id = int(form.get('student_id')) if form.get('student_id') else None
        if not student_id:
            return 'Ученик не выбран', 'error'
        student = Person.query.filter_by(id=student_id).first()
        new_grade = int(form.get('grade')) if form.get('grade') and form.get('grade').isdigit() else None
        new_comment = form.get('comment')
        if not new_comment and not new_grade:
            return 'Нет новой оценки или комментария', 'error'
        record = SchoolLessonJournal.query.filter(
            SchoolLessonJournal.date == record_date,
            SchoolLessonJournal.student_id == student_id,
            SchoolLessonJournal.grade_type == topic,
            SchoolLessonJournal.subject_id == subject.id,
            SchoolLessonJournal.lesson_id == lesson_id,
            SchoolLessonJournal.final_grade == final_grade
        ).first()
        if record:
            record.grade = new_grade
            record.lesson_comment = new_comment
            db.session.flush()
        else:
            new_record = SchoolLessonJournal(
                date=record_date,
                student_id=student_id,
                grade_type=topic,
                subject_id=subject.id,
                lesson_id=lesson_id,
                school_class_id=student.school_class_id,
                final_grade=final_grade,
                grade=new_grade,
                lesson_comment=new_comment
            )
            db.session.add(new_record)
            db.session.flush()

        description = f"Изменение оценки ученика {student.last_name} {student.first_name} по предмету {subject.name} " \
                      f"от {record_date:%d.%m.%Y} ({topic})"
        user_action(user, description)

        return 'Изменения внесены', 'success'


def calc_month_index(date):
    date1 = date.replace(day=1)
    date2 = get_today_date().replace(day=1)

    return relativedelta(date1, date2).months





def student_record(student, month_index):
    result_date = get_today_date() + relativedelta(months=month_index)
    first_date = datetime(result_date.year, result_date.month, 1).date()
    last_date = first_date + relativedelta(months=+1, days=-1)

    student_records = SchoolLessonJournal.query.filter(
        SchoolLessonJournal.date >= first_date,
        SchoolLessonJournal.date <= last_date,
        SchoolLessonJournal.student_id == student.id,
        ~SchoolLessonJournal.final_grade
    ).all()

    student_subjects = Subject.query.filter(
        Subject.subject_type.has(SubjectType.name == "school"),
        Subject.students.any(Person.id == student.id)
    ).order_by(Subject.name).all()
    subjects_dict = {subject.name: {} for subject in student_subjects}
    dates_grade_type_set = set()
    for record in student_records:
        date_str = f'{record.date:%d.%m}'
        grade_type = record.grade_type if record.grade_type else record.lesson.lesson_topic
        dates_grade_type_set.add((date_str, grade_type))
        if record.subject.name in subjects_dict.keys():
            subjects_dict[record.subject.name][(date_str, grade_type)] = {"grade": record.grade,
                                                                          "comment": record.lesson_comment}
        else:
            subjects_dict[record.subject.name] = {(date_str, grade_type): {"grade": record.grade,
                                                                           "comment": record.lesson_comment}}

    dates_grade_type = sorted(list(dates_grade_type_set))

    school_start_year = first_date.year if 9 <= first_date.month <= 12 else first_date.year - 1
    final_grades = SchoolLessonJournal.query.filter(
        SchoolLessonJournal.date >= datetime(school_start_year, 9, 1).date(),
        SchoolLessonJournal.date <= datetime(school_start_year + 1, 7, 1).date(),
        SchoolLessonJournal.student_id == student.id,
        SchoolLessonJournal.final_grade
    ).order_by(SchoolLessonJournal.date).all()

    final_grade_types = []
    for final_grade in final_grades:
        if final_grade.grade_type not in final_grade_types:
            final_grade_types.append(final_grade.grade_type)
        if final_grade.subject.name in subjects_dict.keys():
            subjects_dict[final_grade.subject.name][final_grade.grade_type] = {"grade": final_grade.grade,
                                                                               "comment": final_grade.lesson_comment}
        else:
            subjects_dict[final_grade.subject.name] = {final_grade.grade_type: {"grade": final_grade.grade,
                                                                                "comment": final_grade.lesson_comment}}

    return subjects_dict, dates_grade_type, final_grade_types


def get_period(month_index):
    result_date = get_today_date() + relativedelta(months=month_index)
    return result_date.month, result_date.year


def get_day(day_index):
    result_date = get_today_date() + timedelta(days=day_index)
    return result_date


def calc_day_index(date):
    day_index = (date - get_today_date()).days
    return day_index


def get_after_school_students(period_index, period_type):
    current_period = get_period(0)
    after_school_id = after_school_subject().id
    if period_type == "month":
        date = get_today_date()
        period = get_period(period_index)
    else:
        date = get_day(period_index)
        period = get_period(calc_month_index(date))
    first_date = datetime(period[1], period[0], 1).date()
    last_date = first_date + relativedelta(months=+1, days=-1)
    if period_type == "month":
        after_school_subscriptions = Subscription.query.filter(
            Subscription.subject_id == after_school_id,
            Subscription.purchase_date >= first_date,
            Subscription.purchase_date <= last_date
        ).all()
    else:
        after_school_subscriptions = Subscription.query.filter(
            Subscription.subject_id == after_school_id,
            or_(
                and_(
                    Subscription.period == "month",
                    Subscription.purchase_date >= first_date,
                    Subscription.purchase_date <= last_date
                ),
                and_(
                    Subscription.period == "week",
                    Subscription.purchase_date <= date,
                    Subscription.end_date >= date
                ),
                and_(
                    ~Subscription.period.in_(["month", "week"]),
                    Subscription.purchase_date == date,
                )
            )
        ).all()
    after_school_students = []
    for subscription in after_school_subscriptions:
        after_school_student = subscription.student
        if after_school_student not in after_school_students:
            format_student_info(after_school_student)
            after_school_student.shift = subscription.shift if subscription.shift else 10
            if period == current_period:
                activities = sorted([subject.name for subject in after_school_student.subjects
                                     if subject.subject_type.name not in ["school", "after_school"]])
                after_school_student.activities = activities
            if subscription.period == "month":
                after_school_student.attendance = ["месяц"]
            elif subscription.period == "week":
                after_school_student.attendance = [
                    f"неделя ({subscription.purchase_date:%d.%m}-{subscription.end_date:%d.%m})"
                ]
            elif subscription.period == "day":
                after_school_student.attendance = [f"день ({subscription.purchase_date:%d.%m})"]
            else:
                after_school_student.attendance = [f"{subscription.period} ({subscription.purchase_date:%d.%m})"]
            after_school_students.append(after_school_student)
        else:
            student_ind = after_school_students.index(after_school_student)
            after_school_student = after_school_students[student_ind]
            if subscription.period == "week":
                after_school_student.attendance = [
                    f"неделя ({subscription.purchase_date:%d.%m}-{subscription.end_date:%d.%m})"
                ]
            elif subscription.period == "day":
                after_school_student.attendance.append(f"день ({subscription.purchase_date:%d.%m})")
            else:
                after_school_student.attendance.append(f"{subscription.period} ({subscription.purchase_date:%d.%m})")
    after_school_students = sorted(after_school_students, key=lambda x: (x.shift, x.last_name, x.first_name))

    return after_school_students, period, current_period, date


def get_after_school_prices():
    after_school_prices = SubscriptionType.query.filter(SubscriptionType.period != '').all()
    price_types = []
    for price_type in after_school_prices:
        price_dict = {
            "id": price_type.id,
            "price": float(price_type.price)
        }
        if price_type.period == "месяц":
            price_dict["period"] = "month"
        elif price_type.period == "неделя":
            price_dict["period"] = "week"
        elif price_type.period == "день":
            price_dict["period"] = "day"
        else:
            price_dict["period"] = "hour"
        price_types.append(price_dict)

    return price_types


def handle_after_school_adding(student_id, form, period):
    term = form.get("term")
    shift = int(form.get("shift")) if term in ["month", "day", "week"] and form.get("shift") != "0" else None
    if term == "month":
        purchase_month, purchase_year = period.split("-")
        purchase_date = datetime(int(purchase_year), int(purchase_month), 1).date()
        end_date = purchase_date + relativedelta(months=+1, days=-1)
        period_text = f"{MONTHS[int(purchase_month) - 1]}"
    elif term == "week":
        selected_date = datetime.strptime(form.get("attendance_date"), '%d.%m.%Y').date()
        purchase_date = get_weekday_date(0, selected_date)
        end_date = get_weekday_date(6, selected_date)
        period_text = f"неделя {purchase_date:%d.%m}-{end_date:%d.%m.%Y}"
    else:
        purchase_date = datetime.strptime(form.get("attendance_date"), '%d.%m.%Y').date()
        end_date = purchase_date
        period_text = f"{purchase_date:%d.%m.%Y}"
        if term == "hour":
            hours = int(form.get("hours"))
            term = conjugate_hours(hours)
            period_text = f"{term} {period_text}"

    subscription_type_id = int(form.get("subscription_type"))

    check_after_school = Subscription.query.filter_by(
        subject_id=after_school_subject().id,
        student_id=student_id,
        purchase_date=purchase_date,
        end_date=end_date,
        period=term
    ).first()

    if check_after_school:
        return

    else:
        new_after_school_subscription = Subscription(
            subject_id=after_school_subject().id,
            student_id=student_id,
            subscription_type_id=subscription_type_id,
            purchase_date=purchase_date,
            end_date=end_date,
            shift=shift,
            period=term
        )

        return new_after_school_subscription, period_text


def finance_operation(person, amount, operation_type, description, service, service_id, balance=False, date=None):
    if date is None:
        date = get_today_date()

    if balance:
        person.balance += Decimal(amount)
    elif operation_type == 'balance':
        person.balance += Decimal(amount)
        operation_type = None
        balance = True

    new_operation = Finance(
        person_id=person.id,
        date=date,
        amount=amount,
        operation_type=operation_type,
        student_balance=balance,
        description=description,
        service=service,
        service_id=service_id,
        balance_state=person.balance
    )
    db.session.add(new_operation)
    db.session.flush()


def download_timetable(week):
    workbook = Workbook()
    sheet = workbook.active
    central = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'),
                          top=Side(style='thick'),
                          bottom=Side(style='thick'))
    large_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)

    date_start = get_date(0, week)
    date_end = get_date(4, week)
    dates = get_date_range(week)
    max_length = 1
    last_row_ind = 1

    school_classes = SchoolClass.query.order_by(SchoolClass.school_class).all()
    for school_class in school_classes:
        timetable = Lesson.query.filter(
            Lesson.date >= date_start,
            Lesson.date <= date_end,
            Lesson.school_classes.any(SchoolClass.id == school_class.id)
        ).order_by(Lesson.start_time).all()

        start_end_time = []

        sheet.cell(last_row_ind + 1, 1).value = school_class.school_name
        sheet.cell(last_row_ind + 1, 1).alignment = central
        sheet.cell(last_row_ind + 1, 1).border = thick_border
        sheet.cell(last_row_ind + 1, 1).font = large_font

        for ind, date in enumerate(dates, start=2):
            sheet.cell(last_row_ind, ind).value = DAYS_OF_WEEK[ind - 2]
            sheet.cell(last_row_ind, ind).alignment = central
            sheet.cell(last_row_ind, ind).border = thick_border
            sheet.cell(last_row_ind, ind).font = bold_font
            sheet.cell(last_row_ind + 1, ind).value = date
            sheet.cell(last_row_ind + 1, ind).alignment = central
            sheet.cell(last_row_ind + 1, ind).border = thick_border

        last_row_ind = sheet.max_row

        for lesson in timetable:
            lesson_time = f"{lesson.start_time:%H:%M}-{lesson.end_time:%H:%M}"
            lesson_date = f"{lesson.date:%d.%m}"
            col_ind = dates.index(lesson_date) + 2
            if lesson_time not in start_end_time:
                start_end_time.append(lesson_time)
                ind = len(start_end_time) + last_row_ind
                sheet.cell(ind, 1).value = lesson_time
                sheet.cell(ind, 1).border = thick_border
                sheet.cell(ind, 1).alignment = central
            row_ind = start_end_time.index(lesson_time) + last_row_ind + 1
            cell = sheet.cell(row_ind, col_ind)
            if cell.value:
                cell.value += f" / {lesson.subject.name}"
                first_color = cell.fill.start_color.index
                second_color = lesson.room.color.replace('#', '')
                cell.fill = GradientFill(stop=(first_color, second_color))
            else:
                cell.value = lesson.subject.name
                color = lesson.room.color.replace('#', '')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = central

        new_last_row_ind = sheet.max_row

        for col_ind in range(2, len(dates) + 2):
            for row_ind in range(last_row_ind + 1, new_last_row_ind + 1):
                current_cell = sheet.cell(row_ind, col_ind)
                current_cell.border = thin_border
                if current_cell.value:
                    if len(str(current_cell.value)) > max_length:
                        max_length = len(current_cell.value)

        last_row_ind = new_last_row_ind + 2

    for col in range(1, sheet.max_column + 1):
        adjusted_width = 12 if col == 1 else max_length + 2
        sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

    return workbook, dates


def get_date_range(week):
    date_start = get_date(0, week)
    dates = [f"{date_start + timedelta(day):%d.%m}" for day in range(7)]
    return dates


def filter_day_lessons(form):
    lessons_week = int(form.get('lessons_week'))
    lessons_day = int(form.get('lessons_day'))
    date = get_date(lessons_day, lessons_week)
    query = Lesson.query.filter(Lesson.date == date)
    subject_types = form.get('subject_types')
    if subject_types != "all":
        subject_types_list = [int(subject_type) for subject_type in form.getlist('subject_types_specific')
                              if form.getlist('subject_types_specific')]
        query = query.filter(Lesson.lesson_type_id.in_(subject_types_list))
    lessons = query.all()

    return lessons, date


def del_record(form, record_type, user):
    if record_type == 'student':
        student_id = int(form.get('student_id'))
        student = Person.query.filter_by(id=student_id).first()
        if student:
            student_name = f"{student.last_name} {student.first_name}"
            if student.parents.all():
                for parent in student.parents:
                    if len(parent.children.all()) == 1 and not parent.status and not parent.roles:
                        student.primary_contact = None
                        student.parents.remove(parent)
                        parent.primary_contact = None
                        db.session.delete(parent.contacts[0])
                        db.session.delete(parent)
                db.session.flush()
            if student.children.all() or student.roles:
                student.status = None
                db.session.flush()
            else:
                student.primary_contact = None
                if student.contacts:
                    db.session.delete(student.contacts[0])
                db.session.delete(student)
            db.session.flush()
            user_action(user, f"Удаление клиента {student_name}")
            message = (f"Клиент {student_name} удален", "success")
        else:
            message = ("Такого клиента нет", 'error')

        return message

    if record_type == 'employee':
        employee_id = int(form.get('employee_id'))
        employee = Person.query.filter_by(id=employee_id).first()

        if employee:
            employee_name = f"{employee.last_name} {employee.first_name}"
            future_lessons = []

            if employee.teacher:
                future_lessons = Lesson.query.filter(
                    Lesson.date >= get_today_date(),
                    Lesson.teacher_id == employee_id
                ).all()

            if future_lessons:
                message = (f"Сотрудник {employee_name} не может быть удален", 'error')

            else:
                for role in employee.roles:
                    db.session.delete(role)

                if not employee.children.all() or not employee.status:
                    employee.primary_contact = None
                    db.session.delete(employee.contacts[0])
                    db.session.delete(employee)
                db.session.flush()
                user_action(user, f"Удаление сотрудника {employee_name}")
                message = (f"Сотрудник {employee_name} удален", "success")
        else:
            message = ("Такого сотрудника нет", 'error')

        return message

    if record_type == 'contact':
        contact_person_id = int(form.get('contact_id'))
        student_id = int(form.get('student_id'))
        student = Person.query.filter_by(id=student_id).first()
        contact_person = Person.query.filter_by(id=contact_person_id).first()
        if contact_person:
            if contact_person_id == student_id:
                if contact_person.contacts[0]:
                    db.session.delete(contact_person.contacts[0])
                    db.session.flush()
            else:
                if len(contact_person.children.all()) == 1 and not contact_person.status \
                        and not contact_person.roles:
                    student.parents.remove(contact_person)
                    contact_person.primary_contact = None
                    db.session.delete(contact_person.contacts[0])
                    db.session.delete(contact_person)
                    db.session.flush()
                else:
                    student.parents.remove(contact_person)
                    db.session.flush()

            student_name = f"{student.last_name} {student.first_name}"
            user_action(user, f"Удаление контактной информации клиента {student_name}")
            message = ("Контактная информация удалена", "success")
        else:
            message = ("Такого контакта нет", 'error')

        return message

    if record_type == 'subject':
        subject_id = int(form.get('subject_id'))
        subject = Subject.query.filter(
            Subject.id == subject_id,
            Subject.subject_type.has(~SubjectType.name.in_(['school', 'after_school']))
        ).first()
        if subject:
            subject_name = f"{subject.name} ({subject.subject_type.description})"
            subject_lessons = Lesson.query.filter(subject_id=subject.id).all()
            subscriptions = Subscription.query.filter(
                Subscription.subject_id == subject.id,
                Subscription.end_date >= get_today_date(),
                Subscription.lessons_left > 0
            ).all()
            if subject_lessons or subscriptions:
                text = f"Предмет {subject_name} не может быть удален, "
                text += "т.к. есть дейсвующие абонементы" if subscriptions else "т.к. есть занятия в расписании"
                message = (text, 'error')
            else:
                old_subscriptions = Subscription.query.filter(
                    Subscription.subject_id == subject.id,
                    or_(
                        Subscription.end_date < get_today_date(),
                        Subscription.lessons_left == 0
                    )
                ).all()
                if old_subscriptions:
                    for subscription in old_subscriptions:
                        db.session.delete(subscription)
                db.session.delete(subject)
                db.session.flush()
                user_action(user, f"Удаление предмета {subject_name}")
                message = (f"Предмет {subject_name} удален", 'success')

        else:
            message = ("Такого предмета нет", 'error')

        return message

    if record_type == 'school_subject':
        subject_id = int(form.get('subject_id'))
        subject = Subject.query.filter(
            Subject.id == subject_id,
            Subject.subject_type.has(SubjectType.name == 'school')
        ).first()
        if subject:
            school_class_id = int(form.get('school_class_id'))
            school_class = SchoolClass.query.filter_by(id=school_class_id).first()
            subject.school_classes.remove(school_class)
            subject_name = subject.name
            subject_lessons = Lesson.query.filter_by(subject_id=subject.id).all()
            if not subject_lessons and not subject.school_classes:
                db.session.delete(subject)
                user_action(user, f"Удаление школьного предмета {subject_name}")
                message = (f"Школьный предмет {subject_name} полностью удален", 'success')
            else:
                user_action(user, f"Удаление школьного предмета {subject_name} из класса '{school_class.school_name}'")
                message = (f"Школьный предмет {subject_name} удален из класса", 'success')
            db.session.flush()

        else:
            message = ("Такого предмета нет", 'error')

        return message

    if record_type == 'lesson':
        lesson_id = int(form.get('lesson_id')) if form.get('lesson_id') else None
        lesson = Lesson.query.filter_by(id=lesson_id).first()
        if lesson:
            if not lesson.lesson_completed:
                db.session.delete(lesson)
                db.session.flush()
                user_action(user, f"Отмена занятия {lesson.subject.name} {lesson.date:%d.%m.%Y}")
                return "Занятие отменено", 'success'

            else:
                return "Невозможно отменить занятие", 'error'

        else:
            return "Такого занятия нет", 'error'

    if record_type == 'lessons':
        del_lessons, date = filter_day_lessons(form)
        if del_lessons:
            les = len(del_lessons)
            del_les = 0
            for lesson in del_lessons:
                if not lesson.lesson_completed:
                    db.session.delete(lesson)
                    del_les += 1
                    db.session.flush()

            if del_les == les:
                user_action(user, f"Отмена занятий {date:%d.%m.%Y} ({conjugate_lessons(del_les)})")
                return "Все занятия отменены", 'success'

            elif del_les != 0:
                message1 = (f"{conjugate_lessons(del_les)} отменено", 'success')
                message2 = (f"{conjugate_lessons(del_les)} невозможно отменить", 'error')
                user_action(user, f"Отмена занятий {date:%d.%m.%Y} ({conjugate_lessons(del_les)})")
                return [message1, message2]

            else:
                return "Занятия невозможно отменить", 'error'

        else:
            return "Нет занятий,  удовлетворяющих критериям", 'error'

    if record_type == 'school_student':
        student_id = int(form.get('student_id'))
        student = Person.query.filter(Person.id == student_id, Person.school_class_id).first()
        if student:
            student_name = f"{student.last_name} {student.first_name}"
            school_class = student.school_class.school_name
            student.school_class_id = None
            for subject in student.subjects.all():
                if subject.subject_type.name == "school":
                    student.subjects.remove(subject)
            db.session.flush()
            user_action(user, f"Удаление ученика {student_name} из класса '{school_class}'")
            message = (f"Ученик {student_name} удален из класса", "success")
        else:
            message = ("Такого ученика нет", 'error')

        return message

    if record_type == 'subscription':
        del_subscription_id = int(form.get('subscription_id'))
        del_subscription = Subscription.query.filter_by(id=del_subscription_id).first()
        if del_subscription:
            full_subscription = del_subscription.lessons_left == del_subscription.subscription_type.lessons
            used_subscription = StudentAttendance.query.filter_by(subscription_id=del_subscription_id).all()
            if full_subscription and not used_subscription:
                price = del_subscription.subscription_type.price
                record = Finance.query.filter(
                    Finance.person_id == del_subscription.student_id,
                    Finance.service == "subscription",
                    Finance.service_id == del_subscription.id
                ).first()
                fin_description = f"Возврат за абонемент {del_subscription.subject.name}"
                if record:
                    record.service_id = None
                    finance_operation(del_subscription.student, abs(record.amount), record.operation_type,
                                      fin_description, 'del_subscription', None, balance=record.student_balance)
                else:
                    finance_operation(del_subscription.student, price, 'cash',
                                      fin_description, 'del_subscription', None)

                description = f"Удаление абонемента {del_subscription.subject.name} клиента " \
                              f"{del_subscription.student.last_name} {del_subscription.student.first_name}"
                user_action(user, description)
                db.session.delete(del_subscription)
                db.session.flush()
                message = (f"Абонемент удален", "success")

            else:
                message = (f"Абонемент не может быть удален", "error")

        else:
            message = ("Такого абонемента нет", 'error')

        return message


def change_lessons_date(form):
    change_lessons, old_date = filter_day_lessons(form)
    new_date = datetime.strptime(form.get('new_date'), '%d.%m.%Y').date()
    new_week = calculate_week(new_date)
    if change_lessons:
        les = len(change_lessons)
        change_les = 0
        for lesson in change_lessons:
            if not lesson.lesson_completed:
                start_time = lesson.start_time
                end_time = lesson.end_time
                classes = [cl.id for cl in lesson.school_classes]
                room = lesson.room_id
                teacher = lesson.teacher_id
                split_classes = lesson.split_classes

                conflicting_lessons = check_conflicting_lessons(new_date, start_time, end_time, classes,
                                                                room, teacher, split_classes)
                if not conflicting_lessons:
                    lesson.date = new_date
                    change_les += 1
                    db.session.flush()

        if change_les == les:
            message = [("Все занятия перенесены", 'success')]

        elif change_les != 0:
            message1 = (f"{conjugate_lessons(change_les)} перенесено", 'success')
            message2 = (f"{conjugate_lessons(change_les)} невозможно перенести", 'error')
            message = [message1, message2]

        else:
            new_week = None
            message = [("Занятия невозможно перенести", 'error')]

    else:
        new_week = None
        message = [("Нет занятий,  удовлетворяющих критериям", 'error')]

    return new_week, message, (old_date, new_date)


def add_new_event(form):
    lessons_week = int(form.get('lessons_week'))
    event_day = int(form.get('event_day'))
    event_date = get_date(event_day, lessons_week)
    start_time = datetime.strptime(form.get('event_start_time'), '%H : %M').time()
    end_time = datetime.strptime(form.get('event_end_time'), '%H : %M').time()
    if end_time <= start_time:
        return ('Мероприятие не добавлено. Ошибка во времени проведения', 'error'), None

    room_id = int(form.get('room'))
    conflicts = Lesson.query.filter(
        Lesson.date == event_date,
        Lesson.start_time < end_time,
        Lesson.end_time > start_time,
        Lesson.room_id == room_id
    ).all()
    if not conflicts:
        event_name = form.get('event_name')
        if not event_name:
            return ('Мероприятие не добавлено. Нет названия', 'error'), None

        event_type = SubjectType.query.filter_by(name='event').first()
        existing_event_name = Subject.query.filter_by(name=event_name, subject_type_id=event_type.id).first()
        if existing_event_name:
            event_id = existing_event_name.id
        else:
            new_event_subject = Subject(name=event_name, short_name=event_name, subject_type_id=event_type.id)
            db.session.add(new_event_subject)
            db.session.flush()
            event_id = new_event_subject.id
        new_event = Lesson(
            date=event_date,
            start_time=start_time,
            end_time=end_time,
            lesson_type_id=event_type.id,
            room_id=room_id,
            subject_id=event_id
        )
        db.session.add(new_event)
        db.session.flush()
        return ('Мероприятие добавлено в расписание', 'success'), new_event

    else:
        return ('Мероприятие не добавлено, есть занятия в это же время', 'error'), None


def user_action(user, action_description):
    new_action = UserAction(
        user_id=user.id,
        description=action_description
    )
    db.session.add(new_action)
