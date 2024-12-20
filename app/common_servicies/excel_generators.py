from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, GradientFill, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from app.common_servicies.service import OPERATION_TYPES, get_date, get_date_range, OPERATION_CATEGORIES, DAYS_OF_WEEK
from app.finance.models import Finance
from app.school.models import Person
from app.school.subjects.models import Subject
from app.school_classes.models import SchoolClass
from app.school_classes.service import get_student_school_subjects, format_student_record
from app.timetable.models import StudentAttendance, Lesson

ALIGN_CENTRAL = Alignment(horizontal="center")
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
THICK_BORDER = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))
BOLD_FONT = Font(bold=True)


def sort_finance_operations(report_date):
    lesson_categories = ['Абонемент', 'Занятие']
    finances = Finance.query.filter_by(date=report_date).all()
    day_finance_operations = {oper_type: {} for oper_type in OPERATION_TYPES.keys()}
    day_subjects = set()

    def sort_finances(oper_type, category):
        if category not in day_finance_operations[oper_type].keys():
            day_finance_operations[oper_type][category] = {"Приход": plus, "Расход": minus}
        else:
            day_finance_operations[oper_type][category]["Приход"] += plus
            day_finance_operations[oper_type][category]["Расход"] += minus

    for fin in finances:
        if fin.student_balance:
            plus = fin.amount if fin.amount > 0 else 0
            minus = abs(fin.amount) if fin.amount < 0 else 0
        else:
            plus = abs(fin.amount) if fin.amount < 0 else 0
            minus = fin.amount if fin.amount > 0 else 0
        category = OPERATION_CATEGORIES[fin.service]
        if category in lesson_categories:
            category = fin.subject.name if fin.subject_id else "Занятие"
            day_subjects.add(category)

        if fin.operation_type == "cash":
            sort_finances("cash", category)
        elif fin.operation_type == "bank":
            sort_finances("bank", category)
        if fin.student_balance:
            sort_finances("balance", category)

    subject_list = sorted(list(day_subjects)) if day_subjects else []

    return day_finance_operations, subject_list


def day_completed_lessons(report_date):
    completed_lessons = StudentAttendance.query.join(Subject).filter(
        StudentAttendance.date == report_date
    ).order_by(Subject.name).all()
    day_lessons_dict = {}
    subjects_list = []

    if completed_lessons:
        for record in completed_lessons:
            subject_name = record.subject.name
            subject_type = 'доп' if record.subject.subject_type.name == 'extra' else 'инд'
            subject = (subject_name, subject_type)
            if subject not in subjects_list:
                subjects_list.append(subject)
                day_lessons_dict[subject] = record.price_paid
            else:
                day_lessons_dict[subject] += record.price_paid

    return day_lessons_dict, subjects_list


def download_finance_report(report_date):
    fields = ["Категория", "Приход", "Расход"]
    categories = ['Продленка', 'Депозит, пополнение/возврат', 'Зарплата', 'Питание', 'Школа',
                  'Канцелярия', 'Инкассация', 'Хозтовары', 'Аренда', 'Аттестация', 'Прочее']
    operation_types = OPERATION_TYPES.keys()
    day_finance_operations, subject_categories = sort_finance_operations(report_date)
    all_categories = subject_categories + categories

    completed_lessons_dict, subjects_list = day_completed_lessons(report_date)

    workbook = Workbook()
    sheet = workbook.active
    large_font = Font(bold=True, size=16)
    last_row_ind = 1
    for oper_type in operation_types:
        sheet.merge_cells(f'A{last_row_ind}:C{last_row_ind}')
        sheet[f'A{last_row_ind}'] = OPERATION_TYPES[oper_type]
        sheet[f'A{last_row_ind}'].alignment = ALIGN_CENTRAL
        for row in sheet[f'A{last_row_ind}:C{last_row_ind}']:
            for cell in row:
                cell.border = THICK_BORDER
        sheet[f'A{last_row_ind}'].font = large_font
        last_row_ind += 1

        for ind, field in enumerate(fields, start=1):
            sheet.cell(last_row_ind, ind).value = field
            sheet.cell(last_row_ind, ind).alignment = ALIGN_CENTRAL
            sheet.cell(last_row_ind, ind).border = THIN_BORDER
            sheet.cell(last_row_ind, ind).font = BOLD_FONT
        last_row_ind += 1
        first_row = last_row_ind

        for category in all_categories:
            if category in day_finance_operations[oper_type].keys():
                sheet.cell(last_row_ind, 1).value = category
                sheet.cell(last_row_ind, 1).border = THIN_BORDER
                plus = day_finance_operations[oper_type][category]["Приход"]
                minus = day_finance_operations[oper_type][category]["Расход"]
                sheet.cell(last_row_ind, 2).value = plus
                sheet.cell(last_row_ind, 3).value = minus
                sheet.cell(last_row_ind, 2).border = THIN_BORDER
                sheet.cell(last_row_ind, 3).border = THIN_BORDER
                last_row_ind += 1

        last_row_ind += 1

        if oper_type == "cash":
            sheet.cell(last_row_ind, 1).value = "Остаток на начало дня"
            sheet.cell(last_row_ind, 1).border = THIN_BORDER
            sheet.cell(last_row_ind, 1).font = BOLD_FONT
            sheet.cell(last_row_ind, 2).border = THIN_BORDER
            sheet.cell(last_row_ind, 3).value = 0
            sheet.cell(last_row_ind, 3).font = BOLD_FONT
            sheet.cell(last_row_ind, 3).border = THIN_BORDER
            sheet.cell(last_row_ind + 1, 1).value = "Остаток на конец дня"
            sheet.cell(last_row_ind + 1, 1).border = THIN_BORDER
            sheet.cell(last_row_ind + 1, 1).font = BOLD_FONT
            sheet.cell(last_row_ind + 1, 2).border = THIN_BORDER

            func_str = f'= C{last_row_ind} + SUM(B{first_row}:B{last_row_ind - 2}) - ' \
                       f'SUM(C{first_row}:C{last_row_ind - 2})'
            last_row_ind += 1
            sheet.cell(last_row_ind, 3).value = func_str
            sheet.cell(last_row_ind, 3).border = THIN_BORDER
            sheet.cell(last_row_ind, 3).font = BOLD_FONT

        else:
            sheet.cell(last_row_ind, 1).value = "Оборот за день"
            sheet.cell(last_row_ind, 1).border = THIN_BORDER
            sheet.cell(last_row_ind, 1).font = BOLD_FONT
            sheet.cell(last_row_ind, 2).border = THIN_BORDER
            func_str = f'= SUM(B{first_row}:B{last_row_ind - 2}) - ' \
                       f'SUM(C{first_row}:C{last_row_ind - 2})'
            sheet.cell(last_row_ind, 3).value = func_str
            sheet.cell(last_row_ind, 3).font = BOLD_FONT
            sheet.cell(last_row_ind, 3).border = THIN_BORDER

        last_row_ind += 3

    if subjects_list:
        sheet.merge_cells(f'A{last_row_ind}:C{last_row_ind}')
        sheet[f'A{last_row_ind}'] = f'Занятия {report_date:%d.%m.%y}'
        sheet[f'A{last_row_ind}'].alignment = ALIGN_CENTRAL
        for row in sheet[f'A{last_row_ind}:C{last_row_ind}']:
            for cell in row:
                cell.border = THICK_BORDER
        sheet[f'A{last_row_ind}'].font = large_font
        last_row_ind += 1

        subject_fields = ['Занятие', 'Вид', 'Сумма']
        for ind, field in enumerate(subject_fields, start=1):
            sheet.cell(last_row_ind, ind).value = field
            sheet.cell(last_row_ind, ind).alignment = ALIGN_CENTRAL
            sheet.cell(last_row_ind, ind).border = THIN_BORDER
            sheet.cell(last_row_ind, ind).font = BOLD_FONT
        last_row_ind += 1
        first_subject_row = last_row_ind

        for subject in subjects_list:
            sheet.cell(last_row_ind, 1).value = subject[0]
            sheet.cell(last_row_ind, 1).font = BOLD_FONT
            sheet.cell(last_row_ind, 1).border = THIN_BORDER
            sheet.cell(last_row_ind, 2).value = subject[1]
            sheet.cell(last_row_ind, 2).border = THIN_BORDER
            sheet.cell(last_row_ind, 3).value = float(completed_lessons_dict[subject])
            sheet.cell(last_row_ind, 3).border = THIN_BORDER
            last_row_ind += 1
        subject_func_str = f'= SUM(C{first_subject_row}:C{last_row_ind - 1})'
        sheet.cell(last_row_ind, 1).value = 'Всего'
        sheet.cell(last_row_ind, 1).border = THIN_BORDER
        sheet.cell(last_row_ind, 2).border = THIN_BORDER
        sheet.cell(last_row_ind, 3).value = subject_func_str
        sheet.cell(last_row_ind, 3).border = THIN_BORDER

    for col_ind in range(1, sheet.max_column + 1):
        max_length = 0
        for row_ind in range(1, sheet.max_row + 1):
            current_cell = sheet.cell(row_ind, col_ind)
            if current_cell.value:
                if len(str(current_cell.value)) > max_length and not str(current_cell.value).startswith('='):
                    max_length = len(str(current_cell.value))
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(col_ind)].width = adjusted_width

    return workbook


def download_timetable(week, user):
    workbook = Workbook()
    sheet = workbook.active
    large_font = Font(bold=True, size=16)

    date_start = get_date(0, week)
    date_end = get_date(4, week)
    dates = get_date_range(week)
    max_length = 1
    last_row_ind = 1
    school_classes_query = SchoolClass.query.order_by(SchoolClass.school_class)

    if user.rights == "parent":
        class_list = [person.school_class_id for person in user.user_persons.all()]
        school_classes = school_classes_query.filter(SchoolClass.id.in_(class_list)).all()
    else:
        school_classes = school_classes_query.all()

    for school_class in school_classes:
        timetable = Lesson.query.filter(
            Lesson.date >= date_start,
            Lesson.date <= date_end,
            Lesson.school_classes.any(SchoolClass.id == school_class.id)
        ).order_by(Lesson.start_time).all()

        start_end_time = []

        sheet.cell(last_row_ind + 1, 1).value = school_class.school_name
        sheet.cell(last_row_ind + 1, 1).alignment = ALIGN_CENTRAL
        sheet.cell(last_row_ind + 1, 1).border = THICK_BORDER
        sheet.cell(last_row_ind + 1, 1).font = large_font

        for ind, date in enumerate(dates, start=2):
            sheet.cell(last_row_ind, ind).value = DAYS_OF_WEEK[ind - 2]
            sheet.cell(last_row_ind, ind).alignment = ALIGN_CENTRAL
            sheet.cell(last_row_ind, ind).border = THICK_BORDER
            sheet.cell(last_row_ind, ind).font = BOLD_FONT
            sheet.cell(last_row_ind + 1, ind).value = date
            sheet.cell(last_row_ind + 1, ind).alignment = ALIGN_CENTRAL
            sheet.cell(last_row_ind + 1, ind).border = THICK_BORDER

        last_row_ind = sheet.max_row

        for lesson in timetable:
            lesson_time = f"{lesson.start_time:%H:%M}-{lesson.end_time:%H:%M}"
            lesson_date = f"{lesson.date:%d.%m}"
            col_ind = dates.index(lesson_date) + 2
            if lesson_time not in start_end_time:
                start_end_time.append(lesson_time)
                ind = len(start_end_time) + last_row_ind
                sheet.cell(ind, 1).value = lesson_time
                sheet.cell(ind, 1).border = THICK_BORDER
                sheet.cell(ind, 1).alignment = ALIGN_CENTRAL
            row_ind = start_end_time.index(lesson_time) + last_row_ind + 1
            cell = sheet.cell(row_ind, col_ind)
            if cell.value:
                cell.value += f" / {lesson.subject.name}"
                first_color = cell.fill.start_color.index
                second_color = lesson.room.color.replace('#', '')
                cell.fill = GradientFill(stop=(first_color, second_color))
            else:
                cell.value = lesson.subject.name
                color = lesson.room.color.replace('#', '')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = ALIGN_CENTRAL

        new_last_row_ind = sheet.max_row

        for col_ind in range(2, len(dates) + 2):
            for row_ind in range(last_row_ind + 1, new_last_row_ind + 1):
                current_cell = sheet.cell(row_ind, col_ind)
                current_cell.border = THIN_BORDER
                if current_cell.value:
                    if len(str(current_cell.value)) > max_length:
                        max_length = len(current_cell.value)

        last_row_ind = new_last_row_ind + 2

    for col in range(1, sheet.max_column + 1):
        adjusted_width = 12 if col == 1 else max_length + 2
        sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

    return workbook, dates


def download_school_report(school_class_id, semester, report_date, student=None):
    if student is None:
        school_students = Person.query.filter(
            Person.school_class_id == school_class_id
        ).order_by(Person.last_name, Person.first_name).all()
    else:
        school_students = [student]

    workbook = Workbook()
    thin_thick_border = Border(left=Side(style='thick'),
                               right=Side(style='thick'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))
    col_header_color = PatternFill(start_color='C9DDF0', end_color='C9DDF0', fill_type="solid")
    row_header_color = PatternFill(start_color='FFF5B3', end_color='FFF5B3', fill_type="solid")

    for school_student in school_students:
        sheet = workbook.create_sheet(title=f"{school_student.last_name} {school_student.first_name}")
        subjects_dict = get_student_school_subjects(school_student.id)

        dates_grade_type, subjects_records_dict = format_student_record(
            school_student, semester.start_date, semester.end_date, subjects_dict
        )

        last_row_ind = 1
        sheet.cell(1, 1).value = "Предмет"
        sheet.cell(1, 1).font = BOLD_FONT
        sheet.cell(1, 1).border = THICK_BORDER
        sheet.cell(1, 1).fill = col_header_color
        max_length = len("Предмет")
        for subject, grades_list in subjects_records_dict.items():
            last_row_ind += 1
            last_col_ind = 2
            sheet.cell(last_row_ind, 1).value = subject
            sheet.cell(last_row_ind, 1).font = BOLD_FONT
            sheet.cell(last_row_ind, 1).border = thin_thick_border
            sheet.cell(last_row_ind, 1).fill = row_header_color
            if len(subject) > max_length:
                max_length = len(subject)
            for grade in grades_list.values():
                student_grade = int(grade.get("grade")) if str(grade.get("grade")).isdigit() else None
                if student_grade is not None:
                    comment = None
                    if student_grade <= 2:
                        comment = grade.get("comment")
                    sheet.cell(last_row_ind, last_col_ind).value = student_grade if student_grade != 0 else ''
                    sheet.cell(last_row_ind, last_col_ind).alignment = ALIGN_CENTRAL
                    if comment:
                        cell_comment = Comment(comment, 'Комментарий учителя')
                        sheet.cell(last_row_ind, last_col_ind).comment = cell_comment
                    last_col_ind += 1
        sheet.cell(1, 2).value = "Оценки"
        sheet.cell(1, 2).font = BOLD_FONT
        sheet.cell(1, 2).alignment = ALIGN_CENTRAL
        sheet.column_dimensions['A'].width = max_length + 2
        last_col = sheet.max_column
        if last_col > 2:
            sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
        else:
            sheet.column_dimensions['B'].width = len("Оценки") + 2
        sheet.cell(1, last_col + 1).value = "Средний бал"
        sheet.cell(1, last_col + 1).font = BOLD_FONT
        sheet.cell(1, last_col + 1).alignment = ALIGN_CENTRAL
        sheet.cell(1, last_col + 1).border = THICK_BORDER
        sheet.cell(1, last_col + 1).fill = col_header_color
        sheet.column_dimensions[get_column_letter(last_col + 1)].width = len("Средний бал") + 2
        last_col_letter = get_column_letter(last_col)
        for col in range(2, last_col + 1):
            sheet.cell(1, col).border = THICK_BORDER
            sheet.cell(1, col).fill = col_header_color
            for row in range(2, sheet.max_row + 1):
                if col == last_col:
                    sheet.cell(row, col + 1).border = thin_thick_border
                if row == sheet.max_row:
                    sheet.cell(row, col).border = Border(left=Side(style='thin'),
                                                         right=Side(style='thin'),
                                                         top=Side(style='thin'),
                                                         bottom=Side(style='thick'))
                else:
                    sheet.cell(row, col).border = THIN_BORDER

        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 2).value:
                func_str = f'=ROUND(AVERAGE(B{row}:{last_col_letter}{row}), 2)'
                sheet.cell(row, last_col + 1).value = func_str
            sheet.cell(row, last_col + 1).font = BOLD_FONT
            sheet.cell(row, last_col + 1).alignment = ALIGN_CENTRAL

        for col in [1, last_col + 1]:
            sheet.cell(sheet.max_row, col).border = Border(left=Side(style='thick'),
                                                           right=Side(style='thick'),
                                                           top=Side(style='thin'),
                                                           bottom=Side(style='thick'))
        date_row = sheet.max_row + 2
        sheet.cell(date_row, 1).value = "Дата:"
        sheet.cell(date_row, 1).font = BOLD_FONT
        sheet.merge_cells(start_row=date_row, start_column=2, end_row=date_row, end_column=3)
        sheet.cell(date_row, 2).value = report_date
        sheet.cell(date_row, 2).number_format = 'DD.MM.YYYY'

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    return workbook
